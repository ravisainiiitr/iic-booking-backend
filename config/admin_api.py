"""
Admin-only REST API for frontend admin dashboard.
All endpoints require admin-panel user (admin, manager, operator, finance) or is_staff.
"""
import logging
import os
from datetime import datetime, date, timedelta
from django.conf import settings
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from iic_booking.equipment.image_utils import persist_equipment_image_upload
from django.http import StreamingHttpResponse
from django.utils import timezone
from rest_framework import permissions, status
from rest_framework.exceptions import PermissionDenied
from rest_framework.decorators import action
from rest_framework.pagination import PageNumberPagination
from rest_framework.response import Response
from rest_framework.routers import DefaultRouter
from rest_framework.viewsets import ModelViewSet, ViewSet
from iic_booking.users.models.user_type import UserType
from iic_booking.users.models.department import DepartmentType
from iic_booking.users.rbac import (
    STAFF_ROLE_CODES,
    apply_equipment_department_scope,
    ensure_default_dept_admin_permission_grants,
    get_user_department_scope_id,
    is_department_admin,
    is_main_or_external_relations,
    is_organization_admin,
    scope_queryset_to_department,
    user_has_permission,
)

logger = logging.getLogger(__name__)


class IsAdminPanelUser(permissions.BasePermission):
    """Allow only users with Admin Panel access (Main Admin always; others via role config)."""

    def has_permission(self, request, view):
        if not request.user or not request.user.is_authenticated:
            return False
        from iic_booking.users.rbac import user_has_admin_panel_access

        return user_has_admin_panel_access(request.user)


class IsAdminPanelUserOrReportsStaff(permissions.BasePermission):
    """
    Admin Panel users, or OIC / Lab In-charge with reports.view.

    Lab In-charge and OIC need equipment reports without requiring Admin Panel
    to be enabled for their department role config.
    """

    def has_permission(self, request, view):
        if not request.user or not request.user.is_authenticated:
            return False
        from iic_booking.users.rbac import user_has_admin_panel_access, user_has_permission

        if user_has_admin_panel_access(request.user):
            return True
        ut = getattr(request.user, "user_type", None)
        if ut in {UserType.OPERATOR, UserType.MANAGER}:
            return user_has_permission(
                request.user,
                "reports.view",
                department_id=getattr(request.user, "department_id", None),
            )
        return False


class IsAdminUser(permissions.BasePermission):
    """Allow only admin user type (for Communication management)."""

    def has_permission(self, request, view):
        if not request.user or not request.user.is_authenticated:
            return False
        return getattr(request.user, "user_type", None) == UserType.ADMIN


class IsAdminOrDeptCommunicationAdmin(permissions.BasePermission):
    """Main Admin, or Department Administrator granted admin_settings.communication."""

    def has_permission(self, request, view):
        if not request.user or not request.user.is_authenticated:
            return False
        if getattr(request.user, "is_staff", False):
            return True
        if getattr(request.user, "user_type", None) == UserType.ADMIN:
            return True
        if is_department_admin(request.user) and user_has_permission(
            request.user, "admin_settings.communication", department_id=request.user.department_id
        ):
            return True
        return False


class IsExternalOrgVerifier(permissions.BasePermission):
    """Main Admin or External Relations Administrator (org KYC / verification)."""

    def has_permission(self, request, view):
        if not request.user or not request.user.is_authenticated:
            return False
        if getattr(request.user, "is_staff", False):
            return True
        return is_main_or_external_relations(request.user)


class IsOrgAdminOrAdminPanel(permissions.BasePermission):
    """Admin-panel users, or Organization Administrator for own-org APIs."""

    def has_permission(self, request, view):
        if not request.user or not request.user.is_authenticated:
            return False
        if getattr(request.user, "is_staff", False):
            return True
        user_type = getattr(request.user, "user_type", None)
        if user_type in UserType.get_admin_panel_codes():
            return True
        return is_organization_admin(request.user)


def _request_user_scope_id(request) -> int | None:
    return get_user_department_scope_id(getattr(request, "user", None))


def _require_admin_or_dept_permission(request, code: str, department_id: int | None = None):
    user = getattr(request, "user", None)
    if getattr(user, "user_type", None) == UserType.ADMIN:
        return
    if not user_has_permission(user, code, department_id=department_id):
        raise PermissionDenied("You do not have permission for this departmental action.")


def _require_wallet_manage(request):
    user = getattr(request, "user", None)
    if getattr(user, "user_type", None) == UserType.ADMIN:
        return
    if user_has_permission(user, "wallet.manage") or user_has_permission(user, "admin_settings.wallet"):
        return
    raise PermissionDenied("Wallet management permission is required.")


def _require_bookings_manage(request):
    user = getattr(request, "user", None)
    if getattr(user, "user_type", None) == UserType.ADMIN:
        return
    if not user_has_permission(user, "bookings.manage"):
        raise PermissionDenied("Bookings management permission is required.")


def _require_reports_view(request):
    user = getattr(request, "user", None)
    if getattr(user, "user_type", None) == UserType.ADMIN:
        return
    if not user_has_permission(user, "reports.view"):
        raise PermissionDenied("Reports view permission is required.")

def admin_api_router():
    """Build router for admin-only model APIs. Call from api_router to include under admin/."""
    from django.db.models import Q
    from iic_booking.users.models import (
        Department,
        Project,
        User,
        UserGroup,
        UserGroupMember,
        UserDocument,
        Wallet,
        SubWallet,
        SubWalletTransaction,
        WalletRazorpayOrder,
        WalletRechargeRequest,
        WalletRechargeRequestStatus,
        OrganizationRequest,
        PermissionDefinition,
        DeptAdminPermissionGrant,
        StaffPermissionGrant,
    )
    from iic_booking.users.models.department import DepartmentType, ExternalDepartmentSubcategory, IndianState
    from iic_booking.users.serializers import (
        DepartmentSerializer,
        UserSerializer,
        AdminUserCreateSerializer,
        AdminUserUpdateSerializer,
        AdminUserSetPasswordSerializer,
    )
    from iic_booking.users.serializers.user_serializer import UserBookForListSerializer
    from iic_booking.users.serializers.project_serializer import ProjectSerializer
    from iic_booking.users.serializers.wallet_serializer import (
        WalletSerializer,
        AdminWalletSerializer,
        SubWalletSerializer,
        AdminSubWalletListSerializer,
        AdminSubWalletCreateSerializer,
        SubWalletTransactionSerializer,
        WalletCreditSerializer,
        WalletDebitSerializer,
        WalletRechargeRequestSerializer,
    )
    from iic_booking.equipment.models import (
        Booking,
        DailySlot,
        Equipment,
        EquipmentCategory,
        EquipmentGroup,
        EquipmentGroupQuota,
        Holiday,
        BookingStatus,
        RepeatSampleRequest,
        RepeatSampleRequestStatus,
        SlotStatus,
        CalendarColorSetting,
        BookingChargeSetting,
        InternalUserSlotWindowSetting,
        WaitlistEntry,
        BookingAttemptLog,
        BookingAttemptOutcome,
        EquipmentUserGroup,
        EquipmentUserGroupPurpose,
    )
    from iic_booking.equipment.api_views import refund_booking_internal, get_calendar_colors
    from iic_booking.equipment.serializers import (
        BookingSerializer,
        DailySlotSerializer,
        EquipmentDetailSerializer,
        EquipmentAdminWriteSerializer,
        EquipmentCategorySerializer,
        EquipmentGroupSerializer,
        EquipmentGroupDetailSerializer,
        EquipmentGroupQuotaSerializer,
        RepeatSampleRequestSerializer,
    )
    from rest_framework.parsers import FormParser, JSONParser, MultiPartParser

    from iic_booking.cms.models import MenuItem, HomePageContent, HeroSlide, CmsPage
    from iic_booking.cms.serializers import (
        MenuItemSerializer,
        MenuItemListSerializer,
        HomePageContentSerializer,
        HeroSlideSerializer,
        CmsPageSerializer,
    )

    # Minimal serializers for models that don't have full CRUD serializers yet
    from rest_framework import serializers

    class HolidaySerializer(serializers.ModelSerializer):
        class Meta:
            model = Holiday
            fields = ["id", "date", "reason", "is_active", "color", "created_at", "updated_at"]

    class UserGroupSerializer(serializers.ModelSerializer):
        class Meta:
            model = UserGroup
            fields = ["id", "name", "code", "description", "created_at", "updated_at"]

    class UserGroupMemberSerializer(serializers.ModelSerializer):
        class Meta:
            model = UserGroupMember
            fields = ["id", "user_group", "user", "created_at"]

    class UserDocumentSerializer(serializers.ModelSerializer):
        file_url = serializers.SerializerMethodField()

        class Meta:
            model = UserDocument
            fields = ["id", "user", "file", "file_url", "document_type", "description", "uploaded_at", "updated_at"]
            read_only_fields = ["id", "user", "file", "file_url", "document_type", "description", "uploaded_at", "updated_at"]

        def get_file_url(self, obj):
            if not obj or not obj.file:
                return None
            request = self.context.get("request")
            if request:
                # Serve via backend download endpoint so file is read from configured storage
                # (avoids direct S3/local URL mismatch and NoSuchKey when key is missing)
                path = request.path.rstrip("/")
                if "/user-documents" in path:
                    base = path.split("?")[0]
                    if not base.endswith("/download"):
                        download_path = f"{base}/{obj.pk}/download/"
                        return request.build_absolute_uri(download_path)
                return request.build_absolute_uri(obj.file.url) if obj.file.name else None
            return obj.file.url

    class WalletRazorpayOrderSerializer(serializers.ModelSerializer):
        class Meta:
            model = WalletRazorpayOrder
            fields = ["id", "wallet", "department", "amount_paise", "order_id", "created_at"]
            read_only_fields = ["id", "created_at"]

    class PermissionDefinitionSerializer(serializers.ModelSerializer):
        class Meta:
            model = PermissionDefinition
            fields = ["id", "code", "name", "description"]

    class DeptAdminPermissionGrantSerializer(serializers.ModelSerializer):
        permission_code = serializers.CharField(source="permission.code", read_only=True)
        permission_name = serializers.CharField(source="permission.name", read_only=True)

        class Meta:
            model = DeptAdminPermissionGrant
            fields = [
                "id",
                "department",
                "dept_admin",
                "permission",
                "permission_code",
                "permission_name",
            ]

    class StaffPermissionGrantSerializer(serializers.ModelSerializer):
        permission_code = serializers.CharField(source="permission.code", read_only=True)
        permission_name = serializers.CharField(source="permission.name", read_only=True)

        class Meta:
            model = StaffPermissionGrant
            fields = [
                "id",
                "department",
                "dept_admin",
                "staff_user",
                "permission",
                "permission_code",
                "permission_name",
            ]

    # ViewSets with list/create/retrieve/update/destroy
    class DepartmentViewSet(ModelViewSet):
        permission_classes = [IsAdminPanelUser]
        queryset = Department.objects.all()
        serializer_class = DepartmentSerializer
        parser_classes = [JSONParser, FormParser, MultiPartParser]

        def get_permissions(self):
            if self.action in ("create", "update", "partial_update", "destroy", "bulk_upload_external"):
                return [IsAdminUser()]
            return [IsAdminPanelUser()]

        def get_queryset(self):
            qs = super().get_queryset()
            if is_department_admin(self.request.user):
                return qs.filter(id=self.request.user.department_id)
            return qs

        @action(detail=False, methods=["get"], url_path="bulk-upload-external-template")
        def download_bulk_upload_template(self, request):
            """Download Excel template for bulk upload of external departments."""
            from django.http import HttpResponse
            from openpyxl import Workbook
            from openpyxl.styles import Font

            wb = Workbook()
            ws = wb.active
            ws.title = "External Departments"
            headers = ["Department Name", "State/Union Territory", "Type"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = Font(bold=True)
            # Example row and note
            ws.cell(row=2, column=1, value="Example University")
            ws.cell(row=2, column=2, value="Uttar Pradesh")
            ws.cell(row=2, column=3, value="Educational Institute")
            ws.cell(row=3, column=1, value="Type must be one of: Educational Institute, Govt R&D Organizations, Industry")
            ws.merge_cells("A3:C3")
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = 'attachment; filename="external-departments-template.xlsx"'
            wb.save(response)
            return response

        @action(detail=False, methods=["post"], url_path="bulk-upload-external")
        def bulk_upload_external(self, request):
            """Upload Excel file to create multiple external departments at once."""
            from django.utils.encoding import force_str
            from openpyxl import load_workbook

            f = request.FILES.get("file")
            if not f:
                return Response(
                    {"error": "No file provided. Use form field 'file' with an Excel (.xlsx) file."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            if not f.name.lower().endswith((".xlsx", ".xls")):
                return Response(
                    {"error": "File must be an Excel file (.xlsx)."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            state_choices = dict(IndianState.get_choices())
            state_label_to_value = {force_str(v): k for k, v in state_choices.items()}
            type_choices = dict(ExternalDepartmentSubcategory.get_choices())
            type_label_to_value = {force_str(v): k for k, v in type_choices.items()}

            try:
                wb = load_workbook(f, read_only=True, data_only=True)
            except Exception as e:
                return Response(
                    {"error": f"Invalid Excel file: {str(e)}"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            ws = wb.active
            created = 0
            errors = []

            # Detect header row (first row); expect Department Name, State/Union Territory, Type (any order)
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not header_row:
                return Response(
                    {"error": "Excel sheet is empty."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            header = [str(c).strip().lower() if c else "" for c in header_row]
            col_name = next((i for i, h in enumerate(header) if "department" in h and "name" in h), None)
            col_state = next((i for i, h in enumerate(header) if "state" in h or "union" in h or "territory" in h), None)
            col_type = next((i for i, h in enumerate(header) if h == "type" or (h and "type" in h)), None)
            if col_name is None or col_state is None or col_type is None:
                return Response(
                    {"error": "Expected columns: 'Department Name', 'State/Union Territory', 'Type'."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if row is None:
                    continue
                vals = list(row) if row else []
                name_val = vals[col_name] if col_name < len(vals) else None
                state_val = vals[col_state] if col_state < len(vals) else None
                type_val = vals[col_type] if col_type < len(vals) else None
                name = (str(name_val).strip() if name_val is not None else "") or ""
                state_str = (str(state_val).strip() if state_val is not None else "") or ""
                type_str = (str(type_val).strip() if type_val is not None else "") or ""
                if not name and not state_str and not type_str:
                    continue
                if not name:
                    errors.append({"row": row_idx, "message": "Department name is required."})
                    continue
                if not state_str:
                    errors.append({"row": row_idx, "message": "State/Union Territory is required."})
                    continue
                if not type_str:
                    errors.append({"row": row_idx, "message": "Type is required."})
                    continue
                state_value = state_label_to_value.get(state_str) or (state_str if state_str in state_choices else None)
                type_value = type_label_to_value.get(type_str) or (type_str if type_str in type_choices else None)
                if not state_value:
                    errors.append({"row": row_idx, "message": f"Invalid State/UT: '{state_str}'."})
                    continue
                if not type_value:
                    errors.append({"row": row_idx, "message": f"Invalid Type: '{type_str}'. Use Educational Institute, Govt R&D Organizations, or Industry."})
                    continue
                if Department.objects.filter(name=name).exists():
                    errors.append({"row": row_idx, "message": f"Department with name '{name}' already exists."})
                    continue
                try:
                    Department.objects.create(
                        name=name,
                        department_type=DepartmentType.EXTERNAL,
                        state=state_value,
                        external_subcategory=type_value,
                    )
                    created += 1
                except Exception as e:
                    errors.append({"row": row_idx, "message": str(e)})

            wb.close()
            return Response({
                "created": created,
                "errors": errors,
                "message": f"Created {created} department(s)." + (f" {len(errors)} row(s) had errors." if errors else ""),
            }, status=status.HTTP_200_OK)

    class OrganizationRequestSerializer(serializers.ModelSerializer):
        state_display = serializers.SerializerMethodField()
        status_display = serializers.SerializerMethodField()

        class Meta:
            model = OrganizationRequest
            fields = [
                "id", "name", "approved_name", "state", "state_display",
                "external_subcategory", "email", "requester_name", "web_page", "notes", "status", "status_display",
                "created_department", "approved_by", "created_at", "updated_at",
            ]
            read_only_fields = [
                "id", "state", "external_subcategory", "email", "requester_name", "web_page", "notes",
                "created_department", "approved_by", "created_at", "updated_at",
            ]

        def get_state_display(self, obj):
            if not obj or not obj.state:
                return ""
            from django.utils.encoding import force_str
            return force_str(dict(obj._meta.get_field("state").choices).get(obj.state, obj.state))

        def get_status_display(self, obj):
            if not obj or not obj.status:
                return ""
            from django.utils.encoding import force_str
            return force_str(dict(OrganizationRequest.Status.choices).get(obj.status, obj.status))

    class OrganizationRequestViewSet(ModelViewSet):
        permission_classes = [IsExternalOrgVerifier]
        queryset = OrganizationRequest.objects.all().select_related("created_department", "approved_by").order_by("-created_at")
        serializer_class = OrganizationRequestSerializer
        http_method_names = ["get", "head", "options", "patch", "post"]
        lookup_url_kwarg = "pk"

        def list(self, request, *args, **kwargs):
            """Return paginated organization requests plus standalone external departments (admin-added) as Approved."""
            response = super().list(request, *args, **kwargs)
            # External departments not linked to any organization request (manually added by admin)
            created_dept_ids = set(
                OrganizationRequest.objects.filter(created_department__isnull=False).values_list(
                    "created_department_id", flat=True
                )
            )
            from django.utils.encoding import force_str
            standalone = list(
                Department.objects.filter(department_type=DepartmentType.EXTERNAL)
                .exclude(id__in=created_dept_ids)
                .order_by("name")
            )
            state_choices = dict(Department._meta.get_field("state").choices) if hasattr(Department._meta.get_field("state"), "choices") else {}
            standalone_data = [
                {
                    "id": f"dept-{d.id}",
                    "type": "standalone_department",
                    "department_id": d.id,
                    "name": d.name,
                    "approved_name": d.name,
                    "state": d.state or "",
                    "state_display": force_str(state_choices.get(d.state, d.state or "")),
                    "external_subcategory": d.external_subcategory or "",
                    "email": None,
                    "requester_name": "",
                    "notes": "",
                    "status": "approved",
                    "status_display": "Approved",
                    "created_department": d.id,
                    "approved_by": None,
                    "created_at": d.created_at.isoformat() if d.created_at else None,
                    "updated_at": d.updated_at.isoformat() if d.updated_at else None,
                }
                for d in standalone
            ]
            # When pagination is disabled, response.data is a list; otherwise a dict with "results"
            if isinstance(response.data, list):
                response.data = {"results": response.data, "standalone_departments": standalone_data}
            else:
                response.data["standalone_departments"] = standalone_data
            return response

        def perform_update(self, serializer):
            # Only allow editing name / approved_name
            allowed = {"name", "approved_name"}
            for key in list(serializer.validated_data.keys()):
                if key not in allowed:
                    serializer.validated_data.pop(key, None)
            serializer.save()

        @action(detail=True, methods=["post"], url_path="approve")
        def approve(self, request, pk=None):
            """Create external department from request and approve. Uses approved_name or name as department name."""
            obj = self.get_object()
            if obj.status != OrganizationRequest.Status.PENDING:
                return Response(
                    {"error": f"Request is already {obj.get_status_display()}."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            final_name = (obj.approved_name or obj.name).strip()
            if not final_name:
                return Response({"error": "Organization name is required."}, status=status.HTTP_400_BAD_REQUEST)
            if Department.objects.filter(name=final_name).exists():
                return Response(
                    {"error": f"A department with name '{final_name}' already exists."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            dept = Department.objects.create(
                name=final_name,
                department_type=DepartmentType.EXTERNAL,
                external_subcategory=obj.external_subcategory or ExternalDepartmentSubcategory.GOVT_RND,
                state=obj.state,
            )
            obj.approved_name = final_name
            obj.status = OrganizationRequest.Status.APPROVED
            obj.created_department = dept
            obj.approved_by = request.user
            obj.save(update_fields=["approved_name", "status", "created_department", "approved_by"])
            return Response({"detail": "Approved.", "created_department_id": dept.id})

        @action(detail=True, methods=["post"], url_path="reject")
        def reject(self, request, pk=None):
            obj = self.get_object()
            if obj.status != OrganizationRequest.Status.PENDING:
                return Response(
                    {"error": f"Request is already {obj.get_status_display()}."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            obj.status = OrganizationRequest.Status.REJECTED
            obj.approved_by = request.user
            obj.save(update_fields=["status", "approved_by"])
            return Response({"detail": "Rejected."})

    class ProjectAdminSerializer(ProjectSerializer):
        """Project serializer for admin: allow faculty (FK) for create/update."""
        faculty = serializers.PrimaryKeyRelatedField(queryset=User.objects.all(), required=True)

        class Meta(ProjectSerializer.Meta):
            read_only_fields = ["id", "created_at", "updated_at", "is_expired"]

    class ProjectViewSet(ModelViewSet):
        permission_classes = [IsAdminPanelUser]
        queryset = Project.objects.all().select_related("faculty")
        serializer_class = ProjectAdminSerializer
        lookup_url_kwarg = "pk"

    class UserAdminViewSet(ModelViewSet):
        permission_classes = [IsOrgAdminOrAdminPanel]
        queryset = User.objects.all().select_related("department").order_by("id")
        serializer_class = UserSerializer
        lookup_url_kwarg = "pk"

        def get_permissions(self):
            return [IsOrgAdminOrAdminPanel()]

        def _assert_can_manage_user_payload(self, payload, instance=None):
            user = self.request.user
            if getattr(user, "user_type", None) == UserType.ADMIN:
                return
            if is_organization_admin(user):
                if not user_has_permission(user, "org.users.manage"):
                    raise PermissionDenied("Organization user management permission is required.")
                department_id = payload.get("department", getattr(instance, "department_id", None))
                if department_id in ("", None):
                    department_id = None
                if department_id is None or int(department_id) != int(user.department_id):
                    raise PermissionDenied("Organization Administrators can manage users only inside their own organization.")
                target_user_type = payload.get("user_type", getattr(instance, "user_type", None))
                allowed_org_types = set(UserType.get_external_user_codes()) | {UserType.ORG_ADMIN}
                if target_user_type not in allowed_org_types:
                    raise PermissionDenied("Organization Administrators can only manage external organization users.")
                if target_user_type == UserType.ORG_ADMIN and instance is None:
                    raise PermissionDenied("Cannot create additional Organization Administrators from this panel.")
                return
            if not is_department_admin(user):
                raise PermissionDenied("Only Main Admin or Department Administrator can manage users.")
            _require_admin_or_dept_permission(self.request, "users.manage", department_id=user.department_id)

            department_id = payload.get("department", getattr(instance, "department_id", None))
            if department_id in ("", None):
                department_id = None
            if department_id is None or int(department_id) != int(user.department_id):
                raise PermissionDenied("Department Administrators can manage users only inside their own department.")

            target_user_type = payload.get("user_type", getattr(instance, "user_type", None))
            if target_user_type in {UserType.ADMIN, UserType.DEPT_ADMIN, UserType.EXTERNAL_RELATIONS}:
                raise PermissionDenied("Department Administrators cannot create or edit Main Admin or Department Admin accounts.")
            if target_user_type == UserType.MANAGER and not user_has_permission(user, "oic.assign", department_id=user.department_id):
                raise PermissionDenied("OIC assignment permission is required.")
            if target_user_type == UserType.OPERATOR and not user_has_permission(user, "lab.assign", department_id=user.department_id):
                raise PermissionDenied("Lab In-Charge assignment permission is required.")
            if target_user_type == UserType.FINANCE and not user_has_permission(user, "finance.assign", department_id=user.department_id):
                raise PermissionDenied("Accounts In-Charge assignment permission is required.")

        def get_queryset(self):
            qs = super().get_queryset()
            if is_organization_admin(self.request.user):
                qs = qs.filter(department_id=self.request.user.department_id).exclude(
                    user_type__in=[UserType.ADMIN, UserType.DEPT_ADMIN, UserType.EXTERNAL_RELATIONS]
                )
            else:
                scope_department_id = _request_user_scope_id(self.request)
                if scope_department_id is not None:
                    qs = qs.filter(department_id=scope_department_id)
            if self.action != "list":
                return qs
            search = self.request.query_params.get("search", "").strip()
            if search:
                qs = qs.filter(
                    Q(email__icontains=search) | Q(name__icontains=search)
                )
            restricted = self.request.query_params.get("restricted")
            if restricted is not None and str(restricted).lower() in ("true", "1", "yes"):
                # Default restricted list for manage/section/users:
                # All external users + Govt R&D + Industry + Other + IITR Post Doc/RA/Startups (alias types).
                qs = qs.filter(
                    Q(user_type=UserType.EXTERNAL)
                    | Q(user_type=UserType.RND)
                    | Q(user_type=UserType.INSTITUTE)
                    | Q(user_type=UserType.EXTERNAL_STARTUP_MSME)
                    | Q(user_type=UserType.STARTUP_INCUBATED_IITR)
                    | Q(user_type=UserType.OTHER)
                    | Q(user_type=UserType.ORG_ADMIN)
                    | Q(user_type=UserType.STUDENT, user_type_alias__in=[
                        "IITR Post Doctoral Fellows",
                        "IITR Research Associates in Projects",
                    ])
                    | Q(user_type=UserType.INDIVIDUAL_STUDENT, user_type_alias="IITR Startups")
                )
            user_type = self.request.query_params.get("user_type", "").strip()
            if user_type:
                # Case-insensitive match so "faculty"/"Faculty" both resolve to stored codes.
                by_lower = {str(code).lower(): code for code, _label in UserType.get_choices()}
                canonical = by_lower.get(user_type.lower())
                if canonical:
                    qs = qs.filter(user_type=canonical)
            user_type_alias = self.request.query_params.get("user_type_alias", "").strip()
            if user_type_alias:
                qs = qs.filter(user_type_alias=user_type_alias)
            email_verified = self.request.query_params.get("email_verified")
            if email_verified is not None:
                if str(email_verified).lower() in ("true", "1", "yes"):
                    qs = qs.filter(email_verified=True)
                elif str(email_verified).lower() in ("false", "0", "no"):
                    qs = qs.filter(email_verified=False)
            admin_approved = self.request.query_params.get("admin_approved")
            if admin_approved is not None:
                if str(admin_approved).lower() in ("true", "1", "yes"):
                    qs = qs.filter(admin_approved=True)
                elif str(admin_approved).lower() in ("false", "0", "no"):
                    qs = qs.filter(admin_approved=False)
            is_active = self.request.query_params.get("is_active")
            if is_active is not None:
                if str(is_active).lower() in ("true", "1", "yes"):
                    qs = qs.filter(is_active=True)
                elif str(is_active).lower() in ("false", "0", "no"):
                    qs = qs.filter(is_active=False)
            return qs

        def create(self, request, *args, **kwargs):
            # Dept Admin may create OIC / Lab / Accounts staff in their own department only.
            if is_department_admin(request.user):
                from config.admin_panel_access_api import assert_admin_section_module

                assert_admin_section_module(request.user, "users")
                staff_type = str(request.data.get("user_type") or "").strip().lower()
                allowed = {
                    UserType.MANAGER: "oic.assign",
                    UserType.OPERATOR: "lab.assign",
                    UserType.FINANCE: "finance.assign",
                }
                if staff_type not in allowed:
                    raise PermissionDenied(
                        "Department Administrators can only create Officer In Charge, "
                        "Lab In Charge, or Accounts In Charge users in their department. "
                        "Use Map Channel-i user for existing accounts."
                    )
                _require_admin_or_dept_permission(
                    request, "users.manage", department_id=request.user.department_id
                )
                _require_admin_or_dept_permission(
                    request, allowed[staff_type], department_id=request.user.department_id
                )
                # Force department — ignore client override.
                mutable = request.data.copy() if hasattr(request.data, "copy") else dict(request.data)
                mutable["department"] = request.user.department_id
                mutable["user_type"] = staff_type
                self._assert_can_manage_user_payload(mutable)
                serializer = self.get_serializer(data=mutable)
                serializer.is_valid(raise_exception=True)
                self.perform_create(serializer)
                headers = self.get_success_headers(serializer.data)
                return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

            from config.admin_panel_access_api import assert_admin_section_module

            assert_admin_section_module(request.user, "users")
            self._assert_can_manage_user_payload(request.data)
            return super().create(request, *args, **kwargs)

        def list(self, request, *args, **kwargs):
            from config.admin_panel_access_api import assert_admin_section_module

            assert_admin_section_module(request.user, "users")
            return super().list(request, *args, **kwargs)

        def retrieve(self, request, *args, **kwargs):
            from config.admin_panel_access_api import assert_admin_section_module

            assert_admin_section_module(request.user, "users")
            return super().retrieve(request, *args, **kwargs)

        def get_serializer_class(self):
            if self.action == "create":
                return AdminUserCreateSerializer
            if self.action in ("update", "partial_update"):
                return AdminUserUpdateSerializer
            if self.action == "list" and str(self.request.query_params.get("lite", "")).lower() in (
                "1",
                "true",
                "yes",
            ):
                return UserBookForListSerializer
            return UserSerializer

        def perform_create(self, serializer):
            user = serializer.save()
            ensure_default_dept_admin_permission_grants(user, granted_by=self.request.user)
            if is_organization_admin(self.request.user):
                # Org Admin–created members are approved and active in that organization.
                User.objects.filter(pk=user.pk).update(
                    department_id=self.request.user.department_id,
                    admin_approved=True,
                    email_verified=True,
                    is_active=True,
                    force_inactive=False,
                )
            elif is_department_admin(self.request.user):
                # Dept Admin staff accounts are always in their own department.
                User.objects.filter(pk=user.pk).update(
                    department_id=self.request.user.department_id,
                    admin_approved=True,
                    is_active=True,
                    force_inactive=False,
                )

        def perform_update(self, serializer):
            """Send activation email when user becomes active.

            Business intent: whenever admin action activates a user (including re-activation),
            notify the user via email.
            """
            instance = self.get_object()
            self._assert_can_manage_user_payload(self.request.data, instance=instance)
            was_approved = bool(getattr(instance, "admin_approved", False))
            was_force_inactive = bool(getattr(instance, "force_inactive", False))
            was_active = bool(getattr(instance, "is_active", False))
            send_activation_email = str(self.request.data.get("send_activation_email", "")).lower() in ("1", "true", "yes", "on")
            # Dept Admin cannot move users out of their department.
            if is_department_admin(self.request.user):
                serializer.validated_data["department"] = self.request.user.department
            serializer.save()
            instance.refresh_from_db(fields=["admin_approved", "email_verified", "force_inactive", "is_active", "name", "email"])

            became_active = (not was_active) and bool(instance.is_active)
            became_approved = (not was_approved) and bool(instance.admin_approved)
            no_longer_force_inactive = bool(was_force_inactive) and (not bool(instance.force_inactive))

            # Send activation notification when user becomes active,
            # or when explicitly requested by the UI/admin action.
            # Covers: first-time approval, activate button, re-activation, and "resend activation email".
            should_send = (
                (became_active or send_activation_email)
                and bool(instance.admin_approved)
                and (not bool(instance.force_inactive))
            )
            if should_send:
                try:
                    from iic_booking.communication.service import CommunicationService
                    web_address = (getattr(settings, "FRONTEND_URL", "") or "").strip().rstrip("/") or "/"
                    CommunicationService.send_email(
                        recipient=instance,
                        template="registration_approval_confirmation_email",
                        template_context={
                            "name": instance.name or instance.email,
                            "web_address": web_address,
                        },
                    )
                except Exception as e:
                    logger.exception("Failed to send activation email to %s: %s", instance.email, e)

        def destroy(self, request, *args, **kwargs):
            instance = self.get_object()
            self._assert_can_manage_user_payload(
                {"department": instance.department_id, "user_type": instance.user_type},
                instance=instance,
            )
            return super().destroy(request, *args, **kwargs)

        @action(detail=False, methods=["get"], url_path="mappable-omniport")
        def mappable_omniport(self, request):
            """List same-department Channel-i (Omniport) users that Dept Admin can map to staff roles."""
            actor = request.user
            if is_department_admin(actor):
                ensure_default_dept_admin_permission_grants(actor, granted_by=actor)
            if not (is_department_admin(actor) or getattr(actor, "user_type", None) == UserType.ADMIN):
                raise PermissionDenied("Only Department Administrator or Main Admin can list mappable Omniport users.")
            if is_department_admin(actor):
                department_id = actor.department_id
            else:
                raw = request.query_params.get("department_id")
                try:
                    department_id = int(raw) if raw not in (None, "") else None
                except (TypeError, ValueError):
                    department_id = None
            if department_id is None:
                return Response(
                    {"error": "department_id is required for Main Admin."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            mappable_types = {
                UserType.FACULTY,
                UserType.STUDENT,
                UserType.INDIVIDUAL_STUDENT,
            }
            qs = (
                User.objects.filter(
                    department_id=department_id,
                    user_type__in=mappable_types,
                    is_active=True,
                    force_inactive=False,
                )
                .filter(Q(email_verified=True) | Q(admin_approved=True) | Q(last_login__isnull=False))
                .select_related("department")
                .order_by("name", "email")
            )
            search = (request.query_params.get("search") or "").strip()
            if search:
                qs = qs.filter(Q(email__icontains=search) | Q(name__icontains=search) | Q(emp_id__icontains=search))
            qs = qs[:200]
            data = [
                {
                    "id": u.id,
                    "name": u.name,
                    "email": u.email,
                    "user_type": u.user_type,
                    "user_type_display": u.get_user_type_display_label(),
                    "emp_id": u.emp_id,
                    "department": u.department_id,
                    "department_name": getattr(u.department, "name", None),
                    "email_verified": u.email_verified,
                    "admin_approved": u.admin_approved,
                    "last_login": u.last_login.isoformat() if u.last_login else None,
                }
                for u in qs
            ]
            return Response(data)

        @action(detail=True, methods=["post"], url_path="map-staff-role")
        def map_staff_role(self, request, pk=None):
            """Map an existing Omniport user to OIC / Lab / Accounts within the department."""
            target = self.get_object()
            new_type = (request.data.get("user_type") or "").strip()
            if new_type not in {UserType.MANAGER, UserType.OPERATOR, UserType.FINANCE}:
                return Response(
                    {"error": "user_type must be manager, operator, or finance."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            actor = request.user
            if is_department_admin(actor):
                ensure_default_dept_admin_permission_grants(actor, granted_by=actor)
                if target.department_id != actor.department_id:
                    raise PermissionDenied("You can only map users from your own department.")
                if target.user_type not in {
                    UserType.FACULTY,
                    UserType.STUDENT,
                    UserType.INDIVIDUAL_STUDENT,
                    UserType.MANAGER,
                    UserType.OPERATOR,
                    UserType.FINANCE,
                }:
                    raise PermissionDenied("Only Channel-i (faculty/student) users in your department can be mapped.")
                department_id = actor.department_id
            elif getattr(actor, "user_type", None) == UserType.ADMIN:
                department_id = target.department_id
            else:
                raise PermissionDenied("Only Department Administrator or Main Admin can map staff roles.")
            self._assert_can_manage_user_payload(
                {"department": department_id, "user_type": new_type},
                instance=target,
            )
            target.user_type = new_type
            target.department_id = department_id
            target.admin_approved = True
            target.is_active = True
            target.force_inactive = False
            target.save(
                update_fields=[
                    "user_type",
                    "department_id",
                    "admin_approved",
                    "is_active",
                    "force_inactive",
                ]
            )
            return Response(UserSerializer(target, context={"request": request}).data)

        @action(detail=True, methods=["post"], url_path="set-password")
        def set_password(self, request, pk=None):
            """Set user password (mirrors Django admin /admin/users/user/<id>/password/)."""
            user = self.get_object()
            self._assert_can_manage_user_payload(
                {"department": user.department_id, "user_type": user.user_type},
                instance=user,
            )
            serializer = AdminUserSetPasswordSerializer(
                data=request.data,
                context={"user": user},
            )
            serializer.is_valid(raise_exception=True)
            password = serializer.validated_data["password"]
            user.set_password(password)
            user.save(update_fields=["password"])
            return Response({"detail": "Password has been set."})

    class PermissionDefinitionViewSet(ModelViewSet):
        permission_classes = [IsAdminPanelUser]
        queryset = PermissionDefinition.objects.all().order_by("code")
        serializer_class = PermissionDefinitionSerializer
        http_method_names = ["get"]

        def get_queryset(self):
            from iic_booking.users.rbac import ensure_default_permission_definitions

            ensure_default_permission_definitions()
            return super().get_queryset()

    class DeptAdminPermissionGrantViewSet(ViewSet):
        permission_classes = [IsAdminPanelUser]

        def list(self, request):
            qs = DeptAdminPermissionGrant.objects.select_related("permission", "dept_admin", "department").order_by(
                "dept_admin__email", "permission__code"
            )
            department_id = request.query_params.get("department_id")
            dept_admin_id = request.query_params.get("dept_admin_id")
            if is_department_admin(request.user):
                qs = qs.filter(
                    department_id=request.user.department_id,
                    dept_admin=request.user,
                )
            if department_id:
                qs = qs.filter(department_id=department_id)
            if dept_admin_id:
                qs = qs.filter(dept_admin_id=dept_admin_id)
            serializer = DeptAdminPermissionGrantSerializer(qs, many=True)
            return Response(serializer.data)

        @action(detail=False, methods=["post"], url_path="sync")
        def sync(self, request):
            if getattr(request.user, "user_type", None) != UserType.ADMIN:
                raise PermissionDenied("Only Main Admin can update Department Admin permission caps.")
            dept_admin_id = request.data.get("dept_admin_id")
            permission_codes = request.data.get("permission_codes") or []
            if not dept_admin_id:
                return Response({"error": "dept_admin_id is required."}, status=status.HTTP_400_BAD_REQUEST)
            try:
                dept_admin = User.objects.select_related("department").get(
                    id=int(dept_admin_id),
                    user_type=UserType.DEPT_ADMIN,
                    department__department_type=DepartmentType.INTERNAL,
                )
            except (User.DoesNotExist, ValueError, TypeError):
                return Response({"error": "Department Administrator not found."}, status=status.HTTP_404_NOT_FOUND)
            permissions = list(PermissionDefinition.objects.filter(code__in=permission_codes))
            DeptAdminPermissionGrant.objects.filter(
                dept_admin=dept_admin,
                department_id=dept_admin.department_id,
            ).exclude(permission__in=permissions).delete()
            existing_ids = set(
                DeptAdminPermissionGrant.objects.filter(
                    dept_admin=dept_admin,
                    department_id=dept_admin.department_id,
                ).values_list("permission_id", flat=True)
            )
            DeptAdminPermissionGrant.objects.bulk_create(
                [
                    DeptAdminPermissionGrant(
                        department_id=dept_admin.department_id,
                        dept_admin=dept_admin,
                        permission=permission,
                        granted_by=request.user,
                    )
                    for permission in permissions
                    if permission.id not in existing_ids
                ],
                ignore_conflicts=True,
            )
            serializer = DeptAdminPermissionGrantSerializer(
                DeptAdminPermissionGrant.objects.filter(
                    dept_admin=dept_admin,
                    department_id=dept_admin.department_id,
                ).select_related("permission", "dept_admin", "department"),
                many=True,
            )
            return Response(serializer.data)

    class StaffPermissionGrantViewSet(ViewSet):
        permission_classes = [IsAdminPanelUser]

        def list(self, request):
            qs = StaffPermissionGrant.objects.select_related("permission", "dept_admin", "staff_user", "department").order_by(
                "staff_user__email", "permission__code"
            )
            department_id = request.query_params.get("department_id")
            staff_user_id = request.query_params.get("staff_user_id")
            if is_department_admin(request.user):
                qs = qs.filter(department_id=request.user.department_id)
            if department_id:
                qs = qs.filter(department_id=department_id)
            if staff_user_id:
                qs = qs.filter(staff_user_id=staff_user_id)
            serializer = StaffPermissionGrantSerializer(qs, many=True)
            return Response(serializer.data)

        @action(detail=False, methods=["post"], url_path="sync")
        def sync(self, request):
            actor = request.user
            staff_user_id = request.data.get("staff_user_id")
            permission_codes = request.data.get("permission_codes") or []
            if not staff_user_id:
                return Response({"error": "staff_user_id is required."}, status=status.HTTP_400_BAD_REQUEST)
            try:
                staff_user = User.objects.select_related("department").get(id=int(staff_user_id))
            except (User.DoesNotExist, ValueError, TypeError):
                return Response({"error": "Staff user not found."}, status=status.HTTP_404_NOT_FOUND)

            if getattr(actor, "user_type", None) == UserType.ADMIN:
                dept_admin = User.objects.filter(
                    user_type=UserType.DEPT_ADMIN,
                    department_id=staff_user.department_id,
                ).order_by("id").first()
                if dept_admin is None:
                    return Response({"error": "No Department Administrator is mapped for this department."}, status=status.HTTP_400_BAD_REQUEST)
            else:
                if not is_department_admin(actor):
                    raise PermissionDenied("Only Main Admin or Department Administrator can update staff grants.")
                _require_admin_or_dept_permission(request, "permissions.manage_staff", department_id=actor.department_id)
                if staff_user.department_id != actor.department_id:
                    raise PermissionDenied("Staff grants can only be managed inside your department.")
                if staff_user.user_type not in STAFF_ROLE_CODES:
                    raise PermissionDenied("Only OIC, Lab In-Charge, and Accounts users can receive subordinate grants.")
                dept_admin = actor

            allowed_codes = set(
                DeptAdminPermissionGrant.objects.filter(
                    dept_admin=dept_admin,
                    department_id=dept_admin.department_id,
                ).values_list("permission__code", flat=True)
            )
            requested_codes = [code for code in permission_codes if code in allowed_codes]
            permissions = list(PermissionDefinition.objects.filter(code__in=requested_codes))
            StaffPermissionGrant.objects.filter(
                staff_user=staff_user,
                department_id=dept_admin.department_id,
            ).exclude(permission__in=permissions).delete()
            existing_ids = set(
                StaffPermissionGrant.objects.filter(
                    staff_user=staff_user,
                    department_id=dept_admin.department_id,
                ).values_list("permission_id", flat=True)
            )
            StaffPermissionGrant.objects.bulk_create(
                [
                    StaffPermissionGrant(
                        department_id=dept_admin.department_id,
                        dept_admin=dept_admin,
                        staff_user=staff_user,
                        permission=permission,
                        granted_by=actor,
                    )
                    for permission in permissions
                    if permission.id not in existing_ids
                ],
                ignore_conflicts=True,
            )
            serializer = StaffPermissionGrantSerializer(
                StaffPermissionGrant.objects.filter(
                    staff_user=staff_user,
                    department_id=dept_admin.department_id,
                ).select_related("permission", "dept_admin", "staff_user", "department"),
                many=True,
            )
            return Response(serializer.data)

        @action(detail=True, methods=["get"], url_path="booking-info")
        def booking_info(self, request, pk=None):
            """Return user info needed for admin 'Book slots for user': email, department, Supervisor, balance."""
            user = self.get_object()
            wallet = user.get_accessible_wallet()
            department_name = user.department.name if user.department else None
            wallet_balance = "0.00"
            wallet_faculty_owner = None
            if wallet:
                wallet_balance = str(wallet.total_balance)
                if wallet.user_id != user.id:
                    wallet_faculty_owner = {
                        "name": wallet.user.name or wallet.user.email or "",
                        "email": wallet.user.email or "",
                    }
                else:
                    wallet_faculty_owner = {"name": "Self", "email": user.email or ""}
            return Response({
                "email": user.email or "",
                "department_name": department_name or "",
                "phone_number": getattr(user, "phone_number", None) or "",
                "wallet_faculty_owner": wallet_faculty_owner,
                "wallet_balance": wallet_balance,
            })

        @action(detail=True, methods=["get"], url_path="discounted-charge-equipment")
        def discounted_charge_equipment(self, request, pk=None):
            """
            Return the equipment scope for this user's Discounted Charge Profile.

            Response:
              - use_discounted_charge_profile: bool
              - apply_all_equipment: bool
              - equipment_ids: [int, ...] (empty when apply_all_equipment=true)
            """
            user = self.get_object()
            from iic_booking.equipment.models import UserDiscountedChargeEquipment

            use_discounted = bool(getattr(user, "use_discounted_charge_profile", False))
            if not use_discounted:
                return Response({"use_discounted_charge_profile": False, "apply_all_equipment": False, "equipment_ids": []})

            overrides = UserDiscountedChargeEquipment.objects.filter(user=user, is_active=True).values_list("equipment_id", flat=True)
            ids = list(overrides)
            apply_all = len(ids) == 0
            return Response({
                "use_discounted_charge_profile": True,
                "apply_all_equipment": apply_all,
                "equipment_ids": ids,
            })

        @action(detail=True, methods=["post"], url_path="set-discounted-charge-equipment")
        def set_discounted_charge_equipment(self, request, pk=None):
            """
            Set equipment scope for Discounted Charge Profile.

            Body:
              - use_discounted_charge_profile: bool
              - apply_all_equipment: bool (if true, discounted applies to ALL equipment and overrides are cleared)
              - equipment_ids: [int, ...] (required when apply_all_equipment=false and use_discounted_charge_profile=true)
            """
            user = self.get_object()
            from iic_booking.equipment.models import UserDiscountedChargeEquipment, Equipment

            def _coerce_bool(v):
                if isinstance(v, bool):
                    return v
                if v is None:
                    return False
                return str(v).strip().lower() in ("1", "true", "yes", "on")

            use_discounted = _coerce_bool(request.data.get("use_discounted_charge_profile", False))
            apply_all_equipment = _coerce_bool(request.data.get("apply_all_equipment", True))
            equipment_ids = request.data.get("equipment_ids", [])

            if not isinstance(equipment_ids, list):
                equipment_ids = []

            # Update flag + overrides.
            user.use_discounted_charge_profile = use_discounted
            user.save(update_fields=["use_discounted_charge_profile"])

            # Clear existing overrides first.
            UserDiscountedChargeEquipment.objects.filter(user=user).delete()

            if not use_discounted:
                return Response({"detail": "Discounted charge profile disabled for this user."})

            if apply_all_equipment:
                # Overrides cleared => pricing helper treats it as discounted for all equipment.
                return Response({"detail": "Discounted charge profile set for ALL equipment for this user."})

            equipment_ids = [int(x) for x in equipment_ids if x is not None]
            equipment_ids = [x for x in equipment_ids if x > 0]
            if not equipment_ids:
                return Response(
                    {"error": "equipment_ids must be a non-empty list when apply_all_equipment=false."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # Validate equipments existence.
            existing_ids = set(Equipment.objects.filter(equipment_id__in=equipment_ids).values_list("equipment_id", flat=True))
            missing = [x for x in equipment_ids if x not in existing_ids]
            if missing:
                return Response(
                    {"error": f"Unknown equipment_id(s): {missing[:10]}"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            UserDiscountedChargeEquipment.objects.bulk_create([
                UserDiscountedChargeEquipment(user=user, equipment_id=eid, is_active=True) for eid in existing_ids
            ])
            return Response({"detail": "Discounted charge profile equipment scope updated."})

        @action(detail=True, methods=["get"], url_path="wallet-students")
        def wallet_students(self, request, pk=None):
            """
            For internal faculty: return students (and Other users) who are approved on this faculty's wallet.
            Used by /admin/section/users -> Edit Users popup to bulk-apply discounted charge profile.
            """
            faculty = self.get_object()
            if getattr(faculty, "user_type", None) != UserType.FACULTY:
                return Response({"error": "Only internal faculty users have wallet students."}, status=status.HTTP_400_BAD_REQUEST)

            from iic_booking.users.models.wallet import WalletJoinRequest, WalletJoinRequestStatus

            # Regular students + 'Other' users can join faculty wallets.
            approved_joins = (
                WalletJoinRequest.objects.filter(
                    faculty=faculty,
                    status=WalletJoinRequestStatus.APPROVED,
                    student__user_type__in={UserType.STUDENT, UserType.OTHER},
                )
                .select_related("student")
                .order_by("student__id")
            )

            students_map = {}
            for join in approved_joins:
                st = join.student
                # Defensive: keep only one entry per student.
                students_map[st.id] = {
                    "id": st.id,
                    "email": st.email or "",
                    "name": st.name or "",
                    "user_type": st.user_type or "",
                    "use_discounted_charge_profile": bool(getattr(st, "use_discounted_charge_profile", False)),
                }

            return Response({"students": list(students_map.values())}, status=status.HTTP_200_OK)

        @action(detail=True, methods=["post"], url_path="apply-discounted-charge-profile-to-wallet-students")
        def apply_discounted_charge_profile_to_wallet_students(self, request, pk=None):
            """
            Bulk update use_discounted_charge_profile for students on a faculty wallet.

            Body:
              - apply_all: bool (optional)
              - use_discounted_charge_profile: bool (required unless apply_all provides it)
              - student_updates: [{student_id: int, use_discounted_charge_profile: bool}, ...] (required if apply_all is false)
            """
            faculty = self.get_object()
            if getattr(faculty, "user_type", None) != UserType.FACULTY:
                return Response({"error": "Only internal faculty users have wallet students."}, status=status.HTTP_400_BAD_REQUEST)

            from iic_booking.users.models.wallet import WalletJoinRequest, WalletJoinRequestStatus

            def _coerce_bool(v):
                if isinstance(v, bool):
                    return v
                if v is None:
                    return False
                return str(v).strip().lower() in ("1", "true", "yes", "on")

            apply_all = _coerce_bool(request.data.get("apply_all", False))
            use_discounted_charge_profile = request.data.get("use_discounted_charge_profile", None)
            if use_discounted_charge_profile is not None:
                use_discounted_charge_profile = _coerce_bool(use_discounted_charge_profile)

            apply_all_equipment = _coerce_bool(request.data.get("apply_all_equipment", True))
            equipment_ids = request.data.get("equipment_ids", [])
            if not isinstance(equipment_ids, list):
                equipment_ids = []

            approved_student_ids = list(
                WalletJoinRequest.objects.filter(
                    faculty=faculty,
                    status=WalletJoinRequestStatus.APPROVED,
                    student__user_type__in={UserType.STUDENT, UserType.OTHER},
                ).values_list("student_id", flat=True).distinct()
            )

            if not approved_student_ids:
                return Response({"detail": "No wallet students found.", "updated_count": 0}, status=status.HTTP_200_OK)

            # Build enable/disable sets with validation.
            enable_ids = set()
            disable_ids = set()

            if apply_all:
                if use_discounted_charge_profile is None:
                    return Response({"error": "`use_discounted_charge_profile` is required when `apply_all` is true."}, status=status.HTTP_400_BAD_REQUEST)
                ids = set(approved_student_ids)
                if use_discounted_charge_profile:
                    enable_ids = ids
                else:
                    disable_ids = ids
            else:
                updates = request.data.get("student_updates", None)
                if not isinstance(updates, list) or len(updates) == 0:
                    return Response({"error": "`student_updates` must be a non-empty list when `apply_all` is false."}, status=status.HTTP_400_BAD_REQUEST)

                for item in updates:
                    if not isinstance(item, dict):
                        continue
                    sid = item.get("student_id")
                    if sid is None:
                        continue
                    try:
                        sid = int(sid)
                    except (TypeError, ValueError):
                        continue
                    desired = _coerce_bool(item.get("use_discounted_charge_profile", False))
                    if sid not in approved_student_ids:
                        return Response({"error": f"Student id {sid} is not on this faculty wallet."}, status=status.HTTP_400_BAD_REQUEST)
                    if desired:
                        enable_ids.add(sid)
                    else:
                        disable_ids.add(sid)

            # Apply updates.
            if enable_ids:
                User.objects.filter(id__in=enable_ids).update(use_discounted_charge_profile=True)
                # Clear existing overrides for these users; we'll recreate if needed.
                from iic_booking.equipment.models import UserDiscountedChargeEquipment, Equipment
                UserDiscountedChargeEquipment.objects.filter(user_id__in=list(enable_ids)).delete()

                if not apply_all_equipment:
                    equipment_ids_int = [int(x) for x in equipment_ids if x is not None]
                    equipment_ids_int = [x for x in equipment_ids_int if x > 0]
                    if not equipment_ids_int:
                        return Response(
                            {"error": "`equipment_ids` must be a non-empty list when `apply_all_equipment` is false."},
                            status=status.HTTP_400_BAD_REQUEST,
                        )
                    existing_ids = set(Equipment.objects.filter(equipment_id__in=equipment_ids_int).values_list("equipment_id", flat=True))
                    missing = [x for x in equipment_ids_int if x not in existing_ids]
                    if missing:
                        return Response(
                            {"error": f"Unknown equipment_id(s): {missing[:10]}"},
                            status=status.HTTP_400_BAD_REQUEST,
                        )
                    rows = []
                    for student_id in enable_ids:
                        for eid in existing_ids:
                            rows.append(UserDiscountedChargeEquipment(user_id=student_id, equipment_id=eid, is_active=True))
                    UserDiscountedChargeEquipment.objects.bulk_create(rows)

            if disable_ids:
                User.objects.filter(id__in=disable_ids).update(use_discounted_charge_profile=False)
                from iic_booking.equipment.models import UserDiscountedChargeEquipment
                UserDiscountedChargeEquipment.objects.filter(user_id__in=list(disable_ids)).delete()

            updated_count = len(enable_ids) + len(disable_ids)
            return Response(
                {"detail": "Discounted charge profile updated for wallet students.", "updated_count": updated_count},
                status=status.HTTP_200_OK,
            )

        @action(detail=True, methods=["get"], url_path="transaction-history")
        def transaction_history(self, request, pk=None):
            """Return aggregated sub-wallet transaction history for this user (admin/OIC). Used to verify debit after booking."""
            from iic_booking.users.models import SubWalletTransaction
            from iic_booking.users.repositories.wallet_repository import (
                SubWalletRepository,
                SubWalletTransactionRepository,
            )

            user = self.get_object()
            wallet = user.get_accessible_wallet()
            limit = min(int(request.query_params.get("limit", 50)), 100)
            offset = int(request.query_params.get("offset", 0))

            if not wallet:
                return Response({
                    "user_id": user.id,
                    "user_email": user.email or "",
                    "transactions": [],
                    "count": 0,
                    "limit": limit,
                    "offset": offset,
                }, status=status.HTTP_200_OK)

            sub_wallets = list(SubWalletRepository.get_by_wallet(wallet))
            all_transactions = []
            balance_after_map = {}

            for sub_wallet in sub_wallets:
                txns = list(SubWalletTransactionRepository.get_by_sub_wallet(sub_wallet))
                running = sub_wallet.balance
                for txn in txns:
                    balance_after_map[txn.id] = running
                    if txn.transaction_type == SubWalletTransaction.TransactionType.CREDIT:
                        running -= txn.amount
                    else:
                        running += txn.amount
                    all_transactions.append(txn)

            all_transactions.sort(key=lambda t: t.created_at, reverse=True)
            total_count = len(all_transactions)
            page = all_transactions[offset : offset + limit]

            serializer = SubWalletTransactionSerializer(
                page,
                many=True,
                context={"balance_after_map": balance_after_map},
            )
            return Response({
                "user_id": user.id,
                "user_email": user.email or "",
                "transactions": serializer.data,
                "count": total_count,
                "limit": limit,
                "offset": offset,
            }, status=status.HTTP_200_OK)

    class UserGroupViewSet(ModelViewSet):
        permission_classes = [IsAdminPanelUser]
        queryset = UserGroup.objects.all()
        serializer_class = UserGroupSerializer

    class UserGroupMemberViewSet(ModelViewSet):
        permission_classes = [IsAdminPanelUser]
        queryset = UserGroupMember.objects.all().select_related("user_group", "user")
        serializer_class = UserGroupMemberSerializer

    class UserDocumentViewSet(ModelViewSet):
        http_method_names = ["get", "post", "head", "options"]  # read-only + list
        permission_classes = [IsAdminPanelUser]
        queryset = UserDocument.objects.all().select_related("user")
        serializer_class = UserDocumentSerializer

        def get_queryset(self):
            qs = super().get_queryset()
            if self.action != "list":
                return qs
            user_id = self.request.query_params.get("user", "").strip()
            if user_id:
                try:
                    qs = qs.filter(user_id=int(user_id))
                except ValueError:
                    pass
            return qs.order_by("-uploaded_at")

        @action(detail=True, methods=["get"], url_path="download")
        def download(self, request, pk=None):
            """Stream the document file from configured storage (S3 or local). Fixes direct S3 URL NoSuchKey when file is stored elsewhere."""
            doc = self.get_object()
            if not doc.file:
                return Response({"error": "No file attached."}, status=status.HTTP_404_NOT_FOUND)
            try:
                f = doc.file.open("rb")
            except FileNotFoundError:
                return Response({"error": "File not found in storage."}, status=status.HTTP_404_NOT_FOUND)
            except Exception as e:
                logger.exception("UserDocument download failed for pk=%s (open): %s", pk, e)
                return Response({"error": "Download failed. File may be missing from storage."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

            def stream_and_close(chunk_size=65536):
                try:
                    while True:
                        chunk = f.read(chunk_size)
                        if not chunk:
                            break
                        yield chunk
                finally:
                    f.close()

            filename = os.path.basename(doc.file.name) if doc.file.name else "document"
            response = StreamingHttpResponse(stream_and_close(), content_type="application/octet-stream")
            response["Content-Disposition"] = f'attachment; filename="{filename}"'
            return response

    class WalletViewSet(ModelViewSet):
        permission_classes = [IsAdminPanelUser]
        queryset = Wallet.objects.all().select_related("user")
        serializer_class = AdminWalletSerializer

        def initial(self, request, *args, **kwargs):
            super().initial(request, *args, **kwargs)
            _require_wallet_manage(request)

        def get_queryset(self):
            qs = super().get_queryset()
            if is_department_admin(self.request.user):
                qs = qs.filter(user__department_id=self.request.user.department_id)
            return qs

    class SubWalletViewSet(ModelViewSet):
        permission_classes = [IsAdminPanelUser]
        queryset = SubWallet.objects.all().select_related("wallet", "wallet__user", "department").order_by("department__name")
        serializer_class = SubWalletSerializer
        lookup_url_kwarg = "pk"

        def initial(self, request, *args, **kwargs):
            super().initial(request, *args, **kwargs)
            _require_wallet_manage(request)

        def get_queryset(self):
            qs = super().get_queryset()
            if is_department_admin(self.request.user):
                qs = qs.filter(department_id=self.request.user.department_id)
            return qs

        def get_serializer_class(self):
            if self.action == "create":
                return AdminSubWalletCreateSerializer
            if self.action in ("list", "retrieve"):
                return AdminSubWalletListSerializer
            return SubWalletSerializer

        @action(detail=True, methods=["post"], url_path="credit")
        def credit(self, request, pk=None):
            """Credit the sub-wallet (mirrors Django admin /admin/users/subwallet/<id>/credit/)."""
            sub_wallet = self.get_object()
            serializer = WalletCreditSerializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            amount = serializer.validated_data["amount"]
            description = (serializer.validated_data.get("description") or "").strip() or f"Admin credit"
            try:
                transaction = sub_wallet.credit(amount, description)
            except ValueError as e:
                return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
            return Response({
                "detail": f"Successfully credited ₹{amount} to {sub_wallet.department.name} sub-wallet.",
                "transaction": SubWalletTransactionSerializer(transaction).data,
                "sub_wallet": AdminSubWalletListSerializer(sub_wallet).data,
            })

        @action(detail=True, methods=["post"], url_path="debit")
        def debit(self, request, pk=None):
            """Debit the sub-wallet (mirrors Django admin /admin/users/subwallet/<id>/debit/)."""
            sub_wallet = self.get_object()
            serializer = WalletDebitSerializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            amount = serializer.validated_data["amount"]
            description = (serializer.validated_data.get("description") or "").strip() or f"Admin debit"
            try:
                transaction = sub_wallet.debit(amount, description)
            except ValueError as e:
                return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
            return Response({
                "detail": f"Successfully debited ₹{amount} from {sub_wallet.department.name} sub-wallet.",
                "transaction": SubWalletTransactionSerializer(transaction).data,
                "sub_wallet": AdminSubWalletListSerializer(sub_wallet).data,
            })

    class SubWalletTransactionViewSet(ModelViewSet):
        permission_classes = [IsAdminPanelUser]
        queryset = SubWalletTransaction.objects.all().select_related("sub_wallet")
        serializer_class = SubWalletTransactionSerializer
        http_method_names = ["get", "head", "options", "delete"]

        def initial(self, request, *args, **kwargs):
            super().initial(request, *args, **kwargs)
            _require_wallet_manage(request)

        def get_queryset(self):
            qs = super().get_queryset()
            if is_department_admin(self.request.user):
                qs = qs.filter(sub_wallet__department_id=self.request.user.department_id)
            return qs

        def destroy(self, request, *args, **kwargs):
            """Delete transaction and reverse its effect on sub-wallet balance (mirrors Django admin)."""
            from django.db import transaction as db_transaction
            from django.db.models import F

            txn = self.get_object()
            sub_wallet = txn.sub_wallet
            amount = txn.amount
            txn_id = txn.id
            txn_type = txn.transaction_type
            with db_transaction.atomic():
                if txn_type == SubWalletTransaction.TransactionType.CREDIT:
                    sub_wallet.__class__.objects.filter(pk=sub_wallet.pk).update(balance=F("balance") - amount)
                else:
                    sub_wallet.__class__.objects.filter(pk=sub_wallet.pk).update(balance=F("balance") + amount)
                txn.delete()
            logger.info(
                "SubWalletTransaction %s deleted by user %s (type=%s amount=%s sub_wallet=%s); balance reversed.",
                txn_id,
                getattr(request.user, "id", None),
                txn_type,
                amount,
                getattr(sub_wallet, "id", None),
            )
            return Response(
                {
                    "detail": "Transaction deleted and sub-wallet balance reversed.",
                    "deleted_id": txn_id,
                },
                status=status.HTTP_200_OK,
            )

    class WalletRazorpayOrderViewSet(ModelViewSet):
        permission_classes = [IsAdminPanelUser]
        queryset = WalletRazorpayOrder.objects.all().select_related("wallet", "department")
        serializer_class = WalletRazorpayOrderSerializer

        def initial(self, request, *args, **kwargs):
            super().initial(request, *args, **kwargs)
            _require_wallet_manage(request)

        def get_queryset(self):
            qs = super().get_queryset()
            if is_department_admin(self.request.user):
                qs = qs.filter(department_id=self.request.user.department_id)
            return qs

    class WalletRechargeRequestViewSet(ModelViewSet):
        """Manage internal wallet recharge requests (history, approve/reject/cancel, audit)."""

        permission_classes = [IsAdminPanelUser]
        queryset = (
            WalletRechargeRequest.objects.all()
            .select_related(
                "user",
                "user__department",
                "wallet",
                "department",
                "project",
                "account_incharge",
                "processed_by",
                "fund_receipt_verified_by",
            )
            .prefetch_related("audit_logs", "audit_logs__actor")
            .order_by("-created_at")
        )
        serializer_class = WalletRechargeRequestSerializer
        http_method_names = ["get", "post", "head", "options"]

        class WalletRechargeRequestPagination(PageNumberPagination):
            page_size = 50
            page_size_query_param = "page_size"
            max_page_size = 200

        pagination_class = WalletRechargeRequestPagination

        def initial(self, request, *args, **kwargs):
            super().initial(request, *args, **kwargs)
            _require_wallet_manage(request)

        def get_queryset(self):
            from django.db.models import Q

            qs = super().get_queryset()
            user = self.request.user
            ut = getattr(user, "user_type", None)
            if ut == UserType.ADMIN:
                pass
            elif is_department_admin(user):
                qs = qs.filter(department_id=user.department_id)
            elif ut == UserType.FINANCE:
                qs = qs.filter(
                    Q(account_incharge_id=user.id)
                    | Q(department_id=user.department_id)
                    | Q(department_id__isnull=True)
                )
            else:
                qs = qs.filter(department_id=user.department_id)

            if self.action != "list":
                return qs

            status_filter = (self.request.query_params.get("status") or "").strip().upper()
            if status_filter in {s.value for s in WalletRechargeRequestStatus}:
                qs = qs.filter(status=status_filter)

            search = (self.request.query_params.get("search") or "").strip()
            if search:
                qs = qs.filter(
                    Q(user__email__icontains=search)
                    | Q(user__name__icontains=search)
                    | Q(user__emp_id__icontains=search)
                    | Q(employee_number__icontains=search)
                    | Q(department__name__icontains=search)
                    | Q(department_grant_code__icontains=search)
                    | Q(project_grant_code__icontains=search)
                    | Q(project__project_code__icontains=search)
                    | Q(project__name__icontains=search)
                    | Q(id__icontains=search)
                )

            department_id = (self.request.query_params.get("department") or "").strip()
            if department_id:
                try:
                    qs = qs.filter(department_id=int(department_id))
                except ValueError:
                    pass

            user_id = (self.request.query_params.get("user") or "").strip()
            if user_id:
                try:
                    qs = qs.filter(user_id=int(user_id))
                except ValueError:
                    pass

            project_grant = (self.request.query_params.get("project_grant") or "").strip()
            if project_grant:
                qs = qs.filter(
                    Q(project_grant_code__icontains=project_grant)
                    | Q(project__project_code__icontains=project_grant)
                )

            date_from = (self.request.query_params.get("date_from") or "").strip()
            if date_from:
                try:
                    from datetime import datetime as dt

                    parsed = dt.strptime(date_from, "%Y-%m-%d").date()
                    start_dt = timezone.make_aware(dt.combine(parsed, dt.min.time()))
                    qs = qs.filter(created_at__gte=start_dt)
                except ValueError:
                    pass

            date_to = (self.request.query_params.get("date_to") or "").strip()
            if date_to:
                try:
                    from datetime import datetime as dt

                    parsed = dt.strptime(date_to, "%Y-%m-%d").date()
                    end_dt = timezone.make_aware(dt.combine(parsed + timedelta(days=1), dt.min.time()))
                    qs = qs.filter(created_at__lt=end_dt)
                except ValueError:
                    pass

            ordering = (self.request.query_params.get("ordering") or "-created_at").strip()
            allowed = {
                "created_at",
                "-created_at",
                "amount",
                "-amount",
                "status",
                "-status",
                "responded_at",
                "-responded_at",
            }
            if ordering in allowed:
                qs = qs.order_by(ordering)
            return qs

        @action(detail=True, methods=["post"], url_path="approve")
        def approve(self, request, pk=None):
            from iic_booking.users.models.wallet import WalletRechargeCancellationSource  # noqa: F401
            from iic_booking.users.wallet_recharge_workflow import (
                RechargeAlreadyProcessed,
                already_processed_page,
                approve_request,
                notify_stakeholders_of_decision,
            )

            if getattr(request.user, "user_type", None) == UserType.FINANCE:
                return Response(
                    {
                        "error": (
                            "Department Account In-charge cannot approve recharge requests. "
                            "Use Verify Fund Receipt for financial confirmation."
                        )
                    },
                    status=status.HTTP_403_FORBIDDEN,
                )

            recharge_request = self.get_object()
            if recharge_request.status != WalletRechargeRequestStatus.PENDING:
                page = already_processed_page(
                    recharge_request.status, recharge_request.cancellation_source or ""
                )
                return Response(
                    {
                        "error": page["message"],
                        "page_code": page["page_code"],
                        "already_processed": True,
                        "request": WalletRechargeRequestSerializer(recharge_request).data,
                    },
                    status=status.HTTP_200_OK,
                )
            try:
                approved = approve_request(
                    recharge_request,
                    response_message=(request.data.get("response_message") or "").strip(),
                    actor=request.user,
                    actor_email=request.user.email,
                )
                notify_stakeholders_of_decision(approved)
                return Response(
                    {
                        "message": f"Approved. ₹{approved.amount} credited.",
                        "request": WalletRechargeRequestSerializer(approved).data,
                    },
                    status=status.HTTP_200_OK,
                )
            except RechargeAlreadyProcessed as e:
                recharge_request.refresh_from_db()
                page = already_processed_page(e.status, recharge_request.cancellation_source or "")
                return Response(
                    {"error": page["message"], "page_code": page["page_code"], "already_processed": True},
                    status=status.HTTP_200_OK,
                )
            except ValueError as e:
                return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        @action(detail=True, methods=["post"], url_path="reject")
        def reject(self, request, pk=None):
            from iic_booking.users.models.wallet import WalletRechargeRejectionReason
            from iic_booking.users.wallet_recharge_workflow import (
                RechargeAlreadyProcessed,
                already_processed_page,
                notify_stakeholders_of_decision,
                reject_request,
            )

            if getattr(request.user, "user_type", None) == UserType.FINANCE:
                return Response(
                    {
                        "error": (
                            "Department Account In-charge cannot reject recharge requests. "
                            "Use Verify Fund Receipt for financial confirmation."
                        )
                    },
                    status=status.HTTP_403_FORBIDDEN,
                )

            recharge_request = self.get_object()
            if recharge_request.status != WalletRechargeRequestStatus.PENDING:
                page = already_processed_page(
                    recharge_request.status, recharge_request.cancellation_source or ""
                )
                return Response(
                    {
                        "error": page["message"],
                        "page_code": page["page_code"],
                        "already_processed": True,
                        "request": WalletRechargeRequestSerializer(recharge_request).data,
                    },
                    status=status.HTTP_200_OK,
                )
            reason_code = (
                request.data.get("reason_code") or request.data.get("rejection_reason_code") or ""
            ).strip() or WalletRechargeRejectionReason.OTHER
            reason_text = (
                request.data.get("reason_text")
                or request.data.get("rejection_reason_text")
                or request.data.get("response_message")
                or ""
            ).strip()
            try:
                rejected = reject_request(
                    recharge_request,
                    reason_code=reason_code,
                    reason_text=reason_text,
                    actor=request.user,
                    actor_email=request.user.email,
                )
                notify_stakeholders_of_decision(rejected)
                return Response(
                    {
                        "message": "Wallet recharge request rejected.",
                        "request": WalletRechargeRequestSerializer(rejected).data,
                    },
                    status=status.HTTP_200_OK,
                )
            except RechargeAlreadyProcessed as e:
                recharge_request.refresh_from_db()
                page = already_processed_page(e.status, recharge_request.cancellation_source or "")
                return Response(
                    {"error": page["message"], "page_code": page["page_code"], "already_processed": True},
                    status=status.HTTP_200_OK,
                )
            except ValueError as e:
                return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        @action(detail=True, methods=["post"], url_path="cancel")
        def cancel(self, request, pk=None):
            from iic_booking.users.models.wallet import WalletRechargeCancellationSource
            from iic_booking.users.wallet_recharge_workflow import (
                RechargeAlreadyProcessed,
                already_processed_page,
                cancel_request,
                notify_stakeholders_of_decision,
            )

            if getattr(request.user, "user_type", None) == UserType.FINANCE:
                return Response(
                    {
                        "error": (
                            "Department Account In-charge cannot cancel recharge requests. "
                            "Use Verify Fund Receipt for financial confirmation."
                        )
                    },
                    status=status.HTTP_403_FORBIDDEN,
                )

            recharge_request = self.get_object()
            if recharge_request.status != WalletRechargeRequestStatus.PENDING:
                page = already_processed_page(
                    recharge_request.status, recharge_request.cancellation_source or ""
                )
                return Response(
                    {
                        "error": page["message"],
                        "page_code": page["page_code"],
                        "already_processed": True,
                        "request": WalletRechargeRequestSerializer(recharge_request).data,
                    },
                    status=status.HTTP_200_OK,
                )
            ut = getattr(request.user, "user_type", None)
            source = (
                WalletRechargeCancellationSource.DEPT_ADMIN
                if is_department_admin(request.user)
                else WalletRechargeCancellationSource.ADMIN
            )
            if ut == UserType.FINANCE and not is_department_admin(request.user):
                # Account in-charge cancel is treated as admin cancellation for email status pages
                source = WalletRechargeCancellationSource.ADMIN
            try:
                cancelled = cancel_request(
                    recharge_request,
                    source=source,
                    actor=request.user,
                    actor_email=request.user.email,
                    note=(request.data.get("note") or request.data.get("response_message") or "").strip()
                    or "Cancelled by administrator",
                )
                notify_stakeholders_of_decision(cancelled)
                return Response(
                    {
                        "message": "Wallet recharge request cancelled.",
                        "request": WalletRechargeRequestSerializer(cancelled).data,
                    },
                    status=status.HTTP_200_OK,
                )
            except RechargeAlreadyProcessed as e:
                recharge_request.refresh_from_db()
                page = already_processed_page(e.status, recharge_request.cancellation_source or "")
                return Response(
                    {"error": page["message"], "page_code": page["page_code"], "already_processed": True},
                    status=status.HTTP_200_OK,
                )
            except ValueError as e:
                return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        @action(detail=True, methods=["post"], url_path="verify-fund-receipt")
        def verify_fund_receipt(self, request, pk=None):
            """Department Account In-charge: confirm funds credited to department grant/account."""
            from iic_booking.users.wallet_recharge_workflow import verify_fund_receipt

            ut = getattr(request.user, "user_type", None)
            if ut not in {UserType.FINANCE, UserType.ADMIN} and not is_department_admin(request.user):
                return Response(
                    {"error": "Only Department Account In-charge (or Admin) can verify fund receipt."},
                    status=status.HTTP_403_FORBIDDEN,
                )

            recharge_request = self.get_object()
            remarks = (
                request.data.get("remarks")
                or request.data.get("note")
                or request.data.get("response_message")
                or ""
            ).strip()
            try:
                verified = verify_fund_receipt(
                    recharge_request,
                    actor=request.user,
                    remarks=remarks,
                )
                return Response(
                    {
                        "message": "Fund receipt verified successfully.",
                        "request": WalletRechargeRequestSerializer(verified).data,
                    },
                    status=status.HTTP_200_OK,
                )
            except ValueError as e:
                return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    class BookingViewSet(ModelViewSet):
        permission_classes = [IsAdminPanelUser]
        queryset = Booking.objects.all().select_related("user", "equipment").order_by("-created_at")
        serializer_class = BookingSerializer
        lookup_url_kwarg = "pk"
        lookup_field = "booking_id"

        def initial(self, request, *args, **kwargs):
            super().initial(request, *args, **kwargs)
            _require_bookings_manage(request)

        class BookingPagination(PageNumberPagination):
            page_size = 50

        pagination_class = BookingPagination

        def get_queryset(self):
            qs = super().get_queryset()
            # Non–Main Admin: only bookings for equipment in their department
            qs = scope_queryset_to_department(
                qs, self.request.user, "equipment__internal_department_id"
            )
            if self.action != "list":
                return qs
            status_filter = self.request.query_params.get("status", "").strip().upper()
            if status_filter and status_filter in [s[0] for s in BookingStatus.choices]:
                qs = qs.filter(status=status_filter)
            date_filter = self.request.query_params.get("date", "").strip()
            if date_filter:
                try:
                    from datetime import datetime as dt
                    parsed = dt.strptime(date_filter, "%Y-%m-%d").date()
                    start_dt = timezone.make_aware(dt.combine(parsed, dt.min.time()))
                    end_dt = timezone.make_aware(dt.combine(parsed + timedelta(days=1), dt.min.time()))
                    qs = qs.filter(created_at__gte=start_dt, created_at__lt=end_dt)
                except ValueError:
                    pass
            date_from = self.request.query_params.get("date_from", "").strip()
            if date_from:
                try:
                    from datetime import datetime as dt
                    parsed = dt.strptime(date_from, "%Y-%m-%d").date()
                    start_dt = timezone.make_aware(dt.combine(parsed, dt.min.time()))
                    qs = qs.filter(created_at__gte=start_dt)
                except ValueError:
                    pass
            date_to = self.request.query_params.get("date_to", "").strip()
            if date_to:
                try:
                    from datetime import datetime as dt
                    parsed = dt.strptime(date_to, "%Y-%m-%d").date()
                    end_dt = timezone.make_aware(dt.combine(parsed + timedelta(days=1), dt.min.time()))
                    qs = qs.filter(created_at__lt=end_dt)
                except ValueError:
                    pass
            equipment_id = self.request.query_params.get("equipment", "").strip()
            if equipment_id:
                try:
                    qs = qs.filter(equipment_id=int(equipment_id))
                except ValueError:
                    pass
            return qs

    class RepeatSampleRequestViewSet(ModelViewSet):
        """Read-only list/detail for repeat sample requests; approve/reject via actions (mirrors Django admin /admin/equipment/repeatsamplerequest/)."""
        permission_classes = [IsAdminPanelUser]
        queryset = RepeatSampleRequest.objects.all().select_related(
            "booking", "booking__user", "booking__equipment", "new_booking", "responded_by"
        ).order_by("-requested_at")
        serializer_class = RepeatSampleRequestSerializer
        lookup_url_kwarg = "pk"
        http_method_names = ["get", "post"]  # no put/patch/delete; post only for actions

        class RepeatSampleRequestPagination(PageNumberPagination):
            page_size = 50

        pagination_class = RepeatSampleRequestPagination

        def get_queryset(self):
            from django.db.models import Q
            qs = super().get_queryset()
            qs = scope_queryset_to_department(
                qs, self.request.user, "booking__equipment__internal_department_id"
            )
            if self.action != "list":
                return qs
            status_filter = self.request.query_params.get("status", "").strip().upper()
            if status_filter and status_filter in ("PENDING", "APPROVED", "REJECTED"):
                qs = qs.filter(status=status_filter)
            date_from = self.request.query_params.get("date_from", "").strip()
            if date_from:
                try:
                    from datetime import datetime as dt
                    parsed = dt.strptime(date_from, "%Y-%m-%d").date()
                    start_dt = timezone.make_aware(dt.combine(parsed, dt.min.time()))
                    qs = qs.filter(requested_at__gte=start_dt)
                except ValueError:
                    pass
            date_to = self.request.query_params.get("date_to", "").strip()
            if date_to:
                try:
                    from datetime import datetime as dt
                    parsed = dt.strptime(date_to, "%Y-%m-%d").date()
                    end_dt = timezone.make_aware(dt.combine(parsed + timedelta(days=1), dt.min.time()))
                    qs = qs.filter(requested_at__lt=end_dt)
                except ValueError:
                    pass
            search = self.request.query_params.get("search", "").strip()
            if search:
                qs = qs.filter(
                    Q(booking__virtual_booking_id__icontains=search)
                    | Q(booking__user__email__icontains=search)
                    | Q(booking__equipment__code__icontains=search)
                )
            return qs

        def create(self, request, *args, **kwargs):
            return Response({"detail": "Method not allowed."}, status=status.HTTP_405_METHOD_NOT_ALLOWED)

        def update(self, request, *args, **kwargs):
            return Response({"detail": "Method not allowed."}, status=status.HTTP_405_METHOD_NOT_ALLOWED)

        def partial_update(self, request, *args, **kwargs):
            return Response({"detail": "Method not allowed."}, status=status.HTTP_405_METHOD_NOT_ALLOWED)

        def destroy(self, request, *args, **kwargs):
            return Response({"detail": "Method not allowed."}, status=status.HTTP_405_METHOD_NOT_ALLOWED)

        @action(detail=True, methods=["post"], url_path="approve")
        def approve(self, request, pk=None):
            from iic_booking.equipment.api_views import approve_repeat_sample_request
            return approve_repeat_sample_request(request, int(pk))

        @action(detail=True, methods=["post"], url_path="reject")
        def reject(self, request, pk=None):
            from iic_booking.equipment.api_views import reject_repeat_sample_request
            return reject_repeat_sample_request(request, int(pk))

    class DailySlotViewSet(ModelViewSet):
        permission_classes = [IsAdminPanelUser]
        queryset = DailySlot.objects.all().select_related("slot_master", "booking").order_by("-date", "start_datetime")
        serializer_class = DailySlotSerializer

        class DailySlotPagination(PageNumberPagination):
            page_size = 50

        pagination_class = DailySlotPagination

        def get_queryset(self):
            qs = super().get_queryset()
            qs = scope_queryset_to_department(
                qs, self.request.user, "slot_master__equipment__internal_department_id"
            )
            if self.action != "list":
                return qs
            status_filter = self.request.query_params.get("status", "").strip().upper()
            if status_filter and status_filter in [s[0] for s in SlotStatus.choices]:
                qs = qs.filter(status=status_filter)
            date_filter = self.request.query_params.get("date", "").strip()
            if date_filter:
                try:
                    from datetime import datetime as dt
                    parsed = dt.strptime(date_filter, "%Y-%m-%d").date()
                    qs = qs.filter(date=parsed)
                except ValueError:
                    pass
            date_from = self.request.query_params.get("date_from", "").strip()
            if date_from:
                try:
                    from datetime import datetime as dt
                    parsed = dt.strptime(date_from, "%Y-%m-%d").date()
                    qs = qs.filter(date__gte=parsed)
                except ValueError:
                    pass
            date_to = self.request.query_params.get("date_to", "").strip()
            if date_to:
                try:
                    from datetime import datetime as dt
                    parsed = dt.strptime(date_to, "%Y-%m-%d").date()
                    qs = qs.filter(date__lte=parsed)
                except ValueError:
                    pass
            equipment_id = self.request.query_params.get("equipment", "").strip()
            if equipment_id:
                try:
                    qs = qs.filter(slot_master__equipment_id=int(equipment_id))
                except ValueError:
                    pass
            return qs

        def update(self, request, *args, **kwargs):
            instance = self.get_object()
            new_status = (request.data.get("status") or "").strip().upper()
            slot_status_block = (SlotStatus.BLOCKED, SlotStatus.UNDER_MAINTENANCE, SlotStatus.OPERATOR_ABSENT)
            if new_status in slot_status_block and getattr(instance, "booking_id", None):
                booking = instance.booking
                if booking and str(booking.status) != "REFUNDED":
                    # Operator absent: do NOT auto-refund; apply disruption policy so user can cancel anytime or reschedule.
                    if new_status == SlotStatus.OPERATOR_ABSENT:
                        try:
                            from iic_booking.equipment.maintenance_policy import apply_operator_absent_disruption_for_booking

                            booking = Booking.objects.select_related("equipment", "user").prefetch_related("daily_slots").get(
                                booking_id=booking.booking_id
                            )
                            apply_operator_absent_disruption_for_booking(booking)
                        except Exception as e:
                            return Response(
                                {"error": f"Cannot update slot: disruption policy failed. {str(e)}"},
                                status=status.HTTP_400_BAD_REQUEST,
                            )
                    else:
                        reason_map = {
                            SlotStatus.BLOCKED: "The slot was marked as Blocked.",
                            SlotStatus.UNDER_MAINTENANCE: "The slot was marked as Under Maintenance.",
                            SlotStatus.OPERATOR_ABSENT: "The slot was marked as Operator Absent.",
                        }
                        reason_message = (
                            f"{reason_map.get(new_status, 'The slot is no longer available.')} "
                            "Your booking has been cancelled and a full refund has been issued to your wallet."
                        )
                        try:
                            refund_booking_internal(booking, reason_message, request.user)
                        except Exception as e:
                            return Response(
                                {"error": f"Cannot update slot: refund failed. {str(e)}"},
                                status=status.HTTP_400_BAD_REQUEST,
                            )
            return super().update(request, *args, **kwargs)

    class EquipmentViewSet(ModelViewSet):
        permission_classes = [IsAdminPanelUserOrReportsStaff]
        queryset = Equipment.objects.all().select_related("category", "equipment_group").order_by("code")
        serializer_class = EquipmentDetailSerializer
        lookup_url_kwarg = "pk"
        lookup_field = "equipment_id"

        def initial(self, request, *args, **kwargs):
            super().initial(request, *args, **kwargs)
            from config.admin_panel_access_api import assert_admin_section_module
            from iic_booking.users.rbac import user_has_admin_panel_access, user_has_permission

            # Lab In-charge / OIC with reports.view may list scoped equipment for the Reports filter
            # without full Admin Panel / equipment module access.
            ut = getattr(request.user, "user_type", None)
            if (
                self.action == "list"
                and ut in {UserType.OPERATOR, UserType.MANAGER}
                and user_has_permission(
                    request.user,
                    "reports.view",
                    department_id=getattr(request.user, "department_id", None),
                )
                and not user_has_admin_panel_access(request.user)
            ):
                return
            assert_admin_section_module(request.user, "equipment")

        def _assert_dept_admin_equipment_access(self, payload, instance=None):
            user = self.request.user
            if not is_department_admin(user):
                return
            _require_admin_or_dept_permission(self.request, "equipment.manage", department_id=user.department_id)
            department_id = payload.get("internal_department", getattr(instance, "internal_department_id", None))
            if department_id in ("", None):
                department_id = None
            if department_id is None or int(department_id) != int(user.department_id):
                raise PermissionDenied("Department Administrators can manage equipment only inside their own department.")

        def get_queryset(self):
            from django.db.models import Q
            from iic_booking.equipment.reports import get_equipment_ids_managed_by_oic
            from iic_booking.equipment.api_views import _get_equipment_ids_for_log_access
            qs = super().get_queryset()
            user = self.request.user
            ut = getattr(user, "user_type", None)
            # OIC (manager): only equipment they manage (primary or temporary OIC), within department
            if ut == UserType.MANAGER:
                allowed_ids = get_equipment_ids_managed_by_oic(user.id)
                if not allowed_ids:
                    return qs.none()
                qs = qs.filter(equipment_id__in=allowed_ids)
            # Lab Incharge (operator): only mapped equipment, within department
            elif ut == UserType.OPERATOR:
                allowed_ids = _get_equipment_ids_for_log_access(user) or []
                if not allowed_ids:
                    return qs.none()
                qs = qs.filter(equipment_id__in=allowed_ids)
            # Everyone except Main Admin: strict department isolation
            qs = apply_equipment_department_scope(qs, user)
            if self.action != "list":
                return qs
            search = self.request.query_params.get("search", "").strip()
            if search:
                qs = qs.filter(
                    Q(code__icontains=search)
                    | Q(name__icontains=search)
                    | Q(category__name__icontains=search)
                    | Q(category__code__icontains=search)
                    | Q(equipment_group__name__icontains=search)
                    | Q(equipment_group__code__icontains=search)
                )
            status_filter = self.request.query_params.get("status", "").strip()
            if status_filter:
                qs = qs.filter(status=status_filter)
            profile_type = self.request.query_params.get("profile_type", "").strip()
            if profile_type:
                qs = qs.filter(profile_type=profile_type)
            category_id = self.request.query_params.get("category", "").strip()
            if category_id:
                try:
                    qs = qs.filter(category_id=int(category_id))
                except ValueError:
                    pass
            equipment_group_id = self.request.query_params.get("equipment_group", "").strip()
            if equipment_group_id:
                try:
                    qs = qs.filter(equipment_group_id=int(equipment_group_id))
                except ValueError:
                    pass
            return qs

        def get_serializer_class(self):
            if self.action in ("create", "update", "partial_update"):
                return EquipmentAdminWriteSerializer
            return EquipmentDetailSerializer

        @action(detail=False, methods=["get"], url_path="simple-list")
        def simple_list(self, request):
            """Lightweight equipment list for admin UIs (no pagination)."""
            qs = self.get_queryset().order_by("code").values("equipment_id", "code", "name")
            return Response(list(qs), status=status.HTTP_200_OK)

        def create(self, request, *args, **kwargs):
            if is_department_admin(request.user):
                raise PermissionDenied(
                    "Department Administrators cannot create equipment directly. "
                    "Submit an equipment addition request for Main Admin approval "
                    "(Add Equipment saves as a pending request)."
                )
            self._assert_dept_admin_equipment_access(request.data)
            serializer = self.get_serializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            self.perform_create(serializer)
            instance = serializer.instance
            detail_serializer = EquipmentDetailSerializer(instance, context={"request": request})
            return Response(detail_serializer.data, status=status.HTTP_201_CREATED)

        def update(self, request, *args, **kwargs):
            partial = kwargs.pop("partial", False)
            instance = self.get_object()
            self._assert_dept_admin_equipment_access(request.data, instance=instance)
            serializer = self.get_serializer(instance, data=request.data, partial=partial)
            serializer.is_valid(raise_exception=True)
            self.perform_update(serializer)
            detail_serializer = EquipmentDetailSerializer(instance, context={"request": request})
            return Response(detail_serializer.data)

        @action(detail=True, methods=["post"], url_path="upload-image", parser_classes=[MultiPartParser, FormParser])
        def upload_image(self, request, pk=None):
            equipment = self.get_object()
            file = request.FILES.get("image") or request.FILES.get("file")
            if not file:
                return Response(
                    {"error": "No image file provided. Use form field 'image' or 'file'."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            try:
                persist_equipment_image_upload(equipment, file)
            except Exception as e:
                return Response(
                    {"error": f"Failed to upload image: {str(e)}"},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                )

            detail_serializer = EquipmentDetailSerializer(equipment, context={"request": request})
            return Response(detail_serializer.data)

        @action(detail=True, methods=["post"], url_path="clear-image")
        def clear_image(self, request, pk=None):
            """Clear the equipment catalog image (Django Admin clear checkbox)."""
            equipment = self.get_object()
            old = getattr(equipment, "image", None)
            if old and getattr(old, "name", None):
                try:
                    default_storage.delete(old.name)
                except Exception:
                    pass
            equipment.image = None
            equipment.save(update_fields=["image"])
            detail_serializer = EquipmentDetailSerializer(equipment, context={"request": request})
            return Response(detail_serializer.data)

        @action(detail=True, methods=["post"], url_path="clear-video")
        def clear_video(self, request, pk=None):
            """Clear the equipment video file."""
            equipment = self.get_object()
            old_video = equipment.video_file
            if old_video and old_video.name:
                try:
                    default_storage.delete(old_video.name)
                except Exception:
                    pass
            equipment.video_file = None
            equipment.save(update_fields=["video_file"])
            detail_serializer = EquipmentDetailSerializer(equipment, context={"request": request})
            return Response(detail_serializer.data)

        @action(detail=True, methods=["post"], url_path="upload-video", parser_classes=[MultiPartParser, FormParser])
        def upload_video(self, request, pk=None):
            equipment = self.get_object()
            file = request.FILES.get("video") or request.FILES.get("file")
            if not file:
                return Response(
                    {"error": "No video file provided. Use form field 'video' or 'file'."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            old_video = equipment.video_file
            if old_video and old_video.name:
                try:
                    default_storage.delete(old_video.name)
                except Exception:
                    pass
            try:
                equipment.video_file = file
                equipment.save(update_fields=["video_file"])
            except Exception as e:
                return Response(
                    {"error": f"Failed to upload video: {str(e)}"},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                )
            detail_serializer = EquipmentDetailSerializer(equipment, context={"request": request})
            return Response(detail_serializer.data)

        @action(detail=True, methods=["post"], url_path="bulk-slot-status")
        def bulk_slot_status(self, request, pk=None):
            """
            Set status for slots either by slot_ids or by dates.
            Body: { "slot_ids": [1, 2, ...] } OR { "dates": ["YYYY-MM-DD", ...] } OR { "start_date", "end_date" }
            Plus: "status" (required), "blocked_label" (optional, for BLOCKED).
            Booked slots will be refunded when status is BLOCKED/UNDER_MAINTENANCE/OPERATOR_ABSENT.
            """
            from datetime import datetime as dt, timedelta

            equipment = self.get_object()
            data = request.data or {}
            new_status = (data.get("status") or "").strip().upper()
            valid_statuses = [s[0] for s in SlotStatus.choices]
            if new_status not in valid_statuses:
                return Response(
                    {"error": f"Invalid status. Use one of: {', '.join(valid_statuses)}"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            blocked_label = (data.get("blocked_label") or "").strip() or None
            if new_status != SlotStatus.BLOCKED:
                blocked_label = None

            slot_ids = []
            slot_ids_raw = data.get("slot_ids")
            if slot_ids_raw is not None and isinstance(slot_ids_raw, list):
                for x in slot_ids_raw:
                    try:
                        slot_ids.append(int(x))
                    except (TypeError, ValueError):
                        pass
                slots = list(
                    DailySlot.objects.filter(
                        id__in=slot_ids,
                        slot_master__equipment=equipment,
                    ).select_related("booking"),
                )
                slot_ids = [s.id for s in slots]
            else:
                dates_raw = data.get("dates")
                start_str = (data.get("start_date") or "").strip()
                end_str = (data.get("end_date") or "").strip()
                date_set = set()
                if dates_raw and isinstance(dates_raw, list):
                    for d in dates_raw:
                        s = (d if isinstance(d, str) else str(d)).strip()[:10]
                        try:
                            date_set.add(dt.strptime(s, "%Y-%m-%d").date())
                        except ValueError:
                            pass
                elif start_str and end_str:
                    try:
                        start_d = dt.strptime(start_str[:10], "%Y-%m-%d").date()
                        end_d = dt.strptime(end_str[:10], "%Y-%m-%d").date()
                        if start_d > end_d:
                            start_d, end_d = end_d, start_d
                        d = start_d
                        while d <= end_d:
                            date_set.add(d)
                            d += timedelta(days=1)
                    except ValueError:
                        return Response(
                            {"error": "Invalid start_date or end_date. Use YYYY-MM-DD."},
                            status=status.HTTP_400_BAD_REQUEST,
                        )
                if not date_set:
                    return Response(
                        {"error": "Provide 'slot_ids', 'dates' (list of YYYY-MM-DD), or 'start_date' and 'end_date'."},
                        status=status.HTTP_400_BAD_REQUEST,
                    )

                from iic_booking.equipment.slot_utils import SlotGenerator

                SlotGenerator.ensure_slot_masters_exist(equipment)
                min_date = min(date_set)
                max_date = max(date_set)
                current = min_date
                while current <= max_date:
                    week_end = min(current + timedelta(days=6), max_date)
                    SlotGenerator.generate_slots_for_week(equipment, current, week_end, allow_holiday=True)
                    current = week_end + timedelta(days=1)

                slots = list(
                    DailySlot.objects.filter(
                        slot_master__equipment=equipment,
                        date__in=date_set,
                    ).select_related("booking"),
                )
                slot_ids = [s.id for s in slots]

            if not slot_ids:
                return Response(
                    {"updated": 0, "message": "No slots found for the selection."},
                    status=status.HTTP_200_OK,
                )

            # Booking Not Utilized: only BOOKED slots with a booking are allowed; no refund
            if new_status == SlotStatus.BOOKING_NOT_UTILIZED:
                slots = [s for s in slots if s.status == SlotStatus.BOOKED and s.booking_id]
                slot_ids = [s.id for s in slots]
                if not slot_ids:
                    return Response(
                        {"error": "For 'Booking Not Utilized' only slots that are currently Booked can be selected. No booked slots in your selection."},
                        status=status.HTTP_400_BAD_REQUEST,
                    )

            slot_status_block = (SlotStatus.BLOCKED, SlotStatus.UNDER_MAINTENANCE, SlotStatus.OPERATOR_ABSENT)
            bookings_to_refund = set()
            bookings_for_disruption = set()
            for s in slots:
                if s.booking_id and s.booking and str(getattr(s.booking, "status", "")) != BookingStatus.REFUNDED:
                    if new_status in slot_status_block:
                        if new_status == SlotStatus.OPERATOR_ABSENT:
                            bookings_for_disruption.add(s.booking_id)
                        else:
                            bookings_to_refund.add(s.booking_id)

            reason_map = {
                SlotStatus.BLOCKED: "The slot was marked as Blocked.",
                SlotStatus.UNDER_MAINTENANCE: "The slot was marked as Under Maintenance.",
                SlotStatus.OPERATOR_ABSENT: "The slot was marked as Operator Absent.",
            }
            reason_message = (
                f"{reason_map.get(new_status, 'The slot is no longer available.')} "
                "Your booking has been cancelled and a full refund has been issued to your wallet."
            )
            for booking_id in bookings_to_refund:
                try:
                    booking = Booking.objects.get(booking_id=booking_id)
                    refund_booking_internal(booking, reason_message, request.user)
                except Exception as e:
                    return Response(
                        {"error": f"Cannot update slots: refund failed for booking #{booking_id}. {str(e)}"},
                        status=status.HTTP_400_BAD_REQUEST,
                    )

            if bookings_for_disruption:
                try:
                    from iic_booking.equipment.maintenance_policy import apply_operator_absent_disruption_for_booking
                    qs = (
                        Booking.objects.filter(booking_id__in=list(bookings_for_disruption))
                        .select_related("equipment", "user")
                        .prefetch_related("daily_slots")
                    )
                    for b in qs:
                        apply_operator_absent_disruption_for_booking(b)
                except Exception as e:
                    return Response(
                        {"error": f"Cannot update slots: disruption policy failed. {str(e)}"},
                        status=status.HTTP_400_BAD_REQUEST,
                    )

            update_fields = {"status": new_status, "blocked_label": blocked_label}
            if new_status == SlotStatus.AVAILABLE:
                update_fields["reserved_for_external"] = False
            DailySlot.objects.filter(id__in=slot_ids).update(**update_fields)

            # For Booking Not Utilized: update each affected booking's status so no further changes are allowed
            if new_status == SlotStatus.BOOKING_NOT_UTILIZED:
                from iic_booking.equipment.booking_events import create_booking_event
                from iic_booking.equipment.models import BookingEventType
                seen_bids = set()
                for s in slots:
                    if not s.booking_id or s.booking_id in seen_bids:
                        continue
                    seen_bids.add(s.booking_id)
                    try:
                        b = Booking.objects.get(booking_id=s.booking_id)
                        prev = b.status
                        b.status = BookingStatus.BOOKING_NOT_UTILIZED
                        b.save(update_fields=["status"])
                        create_booking_event(
                            booking=b,
                            event_type=BookingEventType.STATUS_CHANGED,
                            previous_status=prev,
                            new_status=BookingStatus.BOOKING_NOT_UTILIZED,
                            comment="Booking marked as Not Utilized (slot bulk update). No refund issued.",
                            created_by=request.user,
                            send_notification=False,
                        )
                    except Exception as e:
                        logger.warning("Failed to update booking %s status to BOOKING_NOT_UTILIZED: %s", s.booking_id, e)

            # When slots are marked AVAILABLE, notify waitlist so users can book (first come first serve)
            if new_status == SlotStatus.AVAILABLE:
                try:
                    from iic_booking.equipment.waitlist import notify_waitlist_slots_available
                    preferred_slot_ids = [s.id for s in slots]
                    notified = notify_waitlist_slots_available(
                        equipment,
                        preferred_slot_ids=preferred_slot_ids,
                        respect_reschedule_threshold=True,
                    )
                    if notified:
                        logger.info("Notified %d waitlist user(s) for equipment %s (slots available).", notified, equipment.code)
                except Exception as e:
                    logger.warning("Failed to notify waitlist for equipment %s: %s", equipment.code, e)

            # For Booking Not Utilized: send email to each booking user and to Supervisor (no refund)
            if new_status == SlotStatus.BOOKING_NOT_UTILIZED:
                from iic_booking.communication.service import CommunicationService
                from iic_booking.communication.utils import booking_display_id_for_email
                equipment_name = getattr(equipment, "name", "") or getattr(equipment, "code", "") or "Equipment"
                send_to_wallet_owner = data.get("send_email_to_wallet_owner", True)
                if isinstance(send_to_wallet_owner, bool):
                    pass  # use as-is
                else:
                    send_to_wallet_owner = str(send_to_wallet_owner).lower() in ("true", "1", "yes")
                seen_booking_ids = set()
                for s in slots:
                    if not s.booking_id or s.booking_id in seen_booking_ids:
                        continue
                    seen_booking_ids.add(s.booking_id)
                    booking = s.booking
                    if not booking or not booking.user:
                        continue
                    user = booking.user
                    slot_desc = f"{s.date}" + (f" {s.start_datetime.strftime('%H:%M')}-{s.end_datetime.strftime('%H:%M')}" if getattr(s, "start_datetime", None) and getattr(s, "end_datetime", None) else "")
                    ctx = {
                        "user_name": user.name or user.email or "User",
                        "user_email": user.email or "",
                        "equipment_name": equipment_name,
                        "slot_details": slot_desc,
                        "booking_id": booking_display_id_for_email(booking),
                    }
                    try:
                        CommunicationService.send_email(
                            recipient=user,
                            template="booking_not_utilized_email",
                            template_context=ctx,
                            created_by=request.user,
                        )
                    except Exception as e:
                        logger.warning("Failed to send booking not utilized email to %s: %s", user.email, e)
                    # Send to Supervisor if different from user and send_email_to_wallet_owner is True
                    if send_to_wallet_owner:
                        try:
                            wallet = user.get_accessible_wallet()
                            if wallet and wallet.user_id != user.id and wallet.user:
                                owner = wallet.user
                                owner_ctx = {
                                    **ctx,
                                    "student_name": user.name or user.email or "Student",
                                    "student_email": user.email or "",
                                    "wallet_owner_name": owner.name or owner.email or "Faculty",
                                }
                                CommunicationService.send_email(
                                    recipient=owner,
                                    template="booking_not_utilized_wallet_owner_email",
                                    template_context=owner_ctx,
                                    created_by=request.user,
                                )
                        except Exception as e:
                            logger.warning("Failed to send booking not utilized email to Supervisor: %s", e)

            return Response(
                {"updated": len(slot_ids), "message": f"Updated {len(slot_ids)} slot(s) to {new_status}."},
                status=status.HTTP_200_OK,
            )

        @action(detail=True, methods=["post"], url_path="bulk-reserve-external")
        def bulk_reserve_external(self, request, pk=None):
            """
            Mark slots as Reserved for External Users (or unmark).
            Body: { "reserved_for_external": true|false, "slot_ids": [1,2,...] } OR
                  { "reserved_for_external": true|false, "dates": ["YYYY-MM-DD", ...] } OR
                  { "reserved_for_external": true|false, "start_date": "YYYY-MM-DD", "end_date": "YYYY-MM-DD" }.
            Only Admin and OIC can use this. External users can only book slots marked as reserved.
            """
            from datetime import datetime as dt, timedelta
            from iic_booking.equipment.slot_utils import SlotGenerator

            equipment = self.get_object()
            data = request.data or {}
            reserved = data.get("reserved_for_external")
            if reserved is None:
                return Response(
                    {"error": "Missing 'reserved_for_external' (true or false)."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            reserved = reserved in (True, "true", "True", "1", 1)

            slot_ids = []
            slot_ids_raw = data.get("slot_ids")
            if slot_ids_raw is not None and isinstance(slot_ids_raw, list):
                for x in slot_ids_raw:
                    try:
                        slot_ids.append(int(x))
                    except (TypeError, ValueError):
                        pass
                slots = list(
                    DailySlot.objects.filter(
                        id__in=slot_ids,
                        slot_master__equipment=equipment,
                    ),
                )
                slot_ids = [s.id for s in slots]
            else:
                dates_raw = data.get("dates")
                start_str = (data.get("start_date") or "").strip()
                end_str = (data.get("end_date") or "").strip()
                date_set = set()
                if dates_raw and isinstance(dates_raw, list):
                    for d in dates_raw:
                        s = (d if isinstance(d, str) else str(d)).strip()[:10]
                        try:
                            date_set.add(dt.strptime(s, "%Y-%m-%d").date())
                        except ValueError:
                            pass
                elif start_str and end_str:
                    try:
                        start_d = dt.strptime(start_str[:10], "%Y-%m-%d").date()
                        end_d = dt.strptime(end_str[:10], "%Y-%m-%d").date()
                        if start_d > end_d:
                            start_d, end_d = end_d, start_d
                        d = start_d
                        while d <= end_d:
                            date_set.add(d)
                            d += timedelta(days=1)
                    except ValueError:
                        return Response(
                            {"error": "Invalid start_date or end_date. Use YYYY-MM-DD."},
                            status=status.HTTP_400_BAD_REQUEST,
                        )
                if not date_set:
                    return Response(
                        {"error": "Provide 'slot_ids', 'dates' (list of YYYY-MM-DD), or 'start_date' and 'end_date'."},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                SlotGenerator.ensure_slot_masters_exist(equipment)
                min_date = min(date_set)
                max_date = max(date_set)
                current = min_date
                while current <= max_date:
                    week_end = min(current + timedelta(days=6), max_date)
                    SlotGenerator.generate_slots_for_week(equipment, current, week_end, allow_holiday=True)
                    current = week_end + timedelta(days=1)
                slots = list(
                    DailySlot.objects.filter(
                        slot_master__equipment=equipment,
                        date__in=date_set,
                    ),
                )
                slot_ids = [s.id for s in slots]

            if not slot_ids:
                return Response(
                    {"updated": 0, "message": "No slots found for the selection."},
                    status=status.HTTP_200_OK,
                )
            DailySlot.objects.filter(id__in=slot_ids).update(reserved_for_external=reserved)
            label = "Reserved for External Users" if reserved else "Not reserved for external"
            return Response(
                {"updated": len(slot_ids), "message": f"Marked {len(slot_ids)} slot(s) as {label}."},
                status=status.HTTP_200_OK,
            )

        @action(detail=True, methods=["post"], url_path="bulk-home-department-only")
        def bulk_home_department_only(self, request, pk=None):
            """
            Mark slots as reserved for non-home department (or clear the mark).

            Body: { "home_department_only": true|false, "slot_ids": [1,2,...] } OR
                  { "home_department_only": true|false, "dates": ["YYYY-MM-DD", ...] } OR
                  { "home_department_only": true|false, "start_date": "YYYY-MM-DD", "end_date": "YYYY-MM-DD" }.

            Only Admin and OIC.
            true  = reserved for non-home department users
            false = clear reservation (becomes home-department only while any other
                    upcoming reserved mark remains on the equipment; otherwise open to all)
            Unbooked reserved slots open to all departments within Reschedule Hours Threshold
            before the slot start.
            """
            from datetime import datetime as dt, timedelta
            from iic_booking.equipment.slot_utils import SlotGenerator

            equipment = self.get_object()
            data = request.data or {}
            flag = data.get("home_department_only")
            if flag is None:
                return Response(
                    {"error": "Missing 'home_department_only' (true or false)."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            flag = flag in (True, "true", "True", "1", 1)

            slot_ids = []
            slot_ids_raw = data.get("slot_ids")
            if slot_ids_raw is not None and isinstance(slot_ids_raw, list):
                for x in slot_ids_raw:
                    try:
                        slot_ids.append(int(x))
                    except (TypeError, ValueError):
                        pass
                slots = list(
                    DailySlot.objects.filter(
                        id__in=slot_ids,
                        slot_master__equipment=equipment,
                    ),
                )
                slot_ids = [s.id for s in slots]
            else:
                dates_raw = data.get("dates")
                start_str = (data.get("start_date") or "").strip()
                end_str = (data.get("end_date") or "").strip()
                date_set = set()
                if dates_raw and isinstance(dates_raw, list):
                    for d in dates_raw:
                        s = (d if isinstance(d, str) else str(d)).strip()[:10]
                        try:
                            date_set.add(dt.strptime(s, "%Y-%m-%d").date())
                        except ValueError:
                            pass
                elif start_str and end_str:
                    try:
                        start_d = dt.strptime(start_str[:10], "%Y-%m-%d").date()
                        end_d = dt.strptime(end_str[:10], "%Y-%m-%d").date()
                        if start_d > end_d:
                            start_d, end_d = end_d, start_d
                        d = start_d
                        while d <= end_d:
                            date_set.add(d)
                            d += timedelta(days=1)
                    except ValueError:
                        return Response(
                            {"error": "Invalid start_date or end_date. Use YYYY-MM-DD."},
                            status=status.HTTP_400_BAD_REQUEST,
                        )
                if not date_set:
                    return Response(
                        {"error": "Provide 'slot_ids', 'dates' (list of YYYY-MM-DD), or 'start_date' and 'end_date'."},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                SlotGenerator.ensure_slot_masters_exist(equipment)
                min_date = min(date_set)
                max_date = max(date_set)
                current = min_date
                while current <= max_date:
                    week_end = min(current + timedelta(days=6), max_date)
                    SlotGenerator.generate_slots_for_week(equipment, current, week_end, allow_holiday=True)
                    current = week_end + timedelta(days=1)
                slots = list(
                    DailySlot.objects.filter(
                        slot_master__equipment=equipment,
                        date__in=date_set,
                    ),
                )
                slot_ids = [s.id for s in slots]

            if not slot_ids:
                return Response(
                    {"updated": 0, "message": "No slots found for the selection."},
                    status=status.HTTP_200_OK,
                )
            DailySlot.objects.filter(id__in=slot_ids).update(home_department_only=flag)
            label = (
                "Reserved for non-home department"
                if flag
                else "Home department (cleared non-home reservation)"
            )
            return Response(
                {"updated": len(slot_ids), "message": f"Marked {len(slot_ids)} slot(s) as {label}."},
                status=status.HTTP_200_OK,
            )

        @action(detail=True, methods=["get"], url_path="waitlist")
        def waitlist(self, request, pk=None):
            """List waitlist entries for this equipment (admin/OIC). Ordered by created_at (FIFO)."""
            equipment = self.get_object()
            entries = (
                WaitlistEntry.objects.filter(equipment=equipment)
                .select_related("user")
                .order_by("created_at")
            )
            active_count = WaitlistEntry.objects.filter(equipment=equipment, status="ACTIVE").count()
            cannot_fulfill_count = WaitlistEntry.objects.filter(equipment=equipment, status="CANNOT_FULFILL").count()
            user_ids = [e.user_id for e in entries if getattr(e, "user_id", None) is not None]
            failed_logs_by_user: dict = {}
            if user_ids:
                logs = (
                    BookingAttemptLog.objects.filter(
                        equipment=equipment,
                        user_id__in=user_ids,
                        outcome=BookingAttemptOutcome.FAILED,
                    )
                    .order_by("user_id", "-requested_at")
                )
                for lg in logs:
                    failed_logs_by_user.setdefault(lg.user_id, []).append(lg)
            position = 0
            result = []
            for e in entries:
                position += 1
                user_logs = failed_logs_by_user.get(getattr(e, "user_id", None), [])
                # Prefer a log that happened before the waitlist entry was created.
                matched_log = None
                if user_logs and getattr(e, "created_at", None) is not None:
                    for lg in user_logs:
                        if getattr(lg, "requested_at", None) is not None and lg.requested_at <= e.created_at:
                            matched_log = lg
                            break
                if matched_log is None and user_logs:
                    matched_log = user_logs[0]

                failure_reason = getattr(matched_log, "failure_reason", "") if matched_log else ""
                requested_at = getattr(matched_log, "requested_at", None) if matched_log else None
                number_of_samples = getattr(matched_log, "number_of_samples", None) if matched_log else None
                slots_requested = getattr(matched_log, "slots_requested", None) if matched_log else None
                duration_minutes = getattr(matched_log, "duration_minutes", None) if matched_log else None
                additional_info = getattr(matched_log, "additional_info", None) if matched_log else None

                result.append({
                    "id": e.id,
                    "position": position,
                    "user_id": e.user_id,
                    "user_email": getattr(e.user, "email", ""),
                    "user_name": getattr(e.user, "name", None) or getattr(e.user, "email", ""),
                    "created_at": e.created_at.isoformat() if e.created_at else None,
                    "status": getattr(e, "status", None) or "ACTIVE",
                    "cannot_fulfill_remark": getattr(e, "cannot_fulfill_remark", None),
                    "marked_cannot_fulfill_at": (
                        e.marked_cannot_fulfill_at.isoformat()
                        if getattr(e, "marked_cannot_fulfill_at", None)
                        else None
                    ),
                    "booking_attempt_requested_at": requested_at.isoformat() if requested_at else None,
                    "booking_attempt_failure_reason": failure_reason or "",
                    "booking_attempt_number_of_samples": number_of_samples,
                    "booking_attempt_slots_requested": slots_requested,
                    "booking_attempt_duration_minutes": duration_minutes,
                    "booking_attempt_additional_info": additional_info,
                })
            return Response({
                "equipment_id": equipment.equipment_id,
                "equipment_code": equipment.code,
                "equipment_name": getattr(equipment, "name", "") or equipment.code,
                "waitlist_queue_depth": getattr(equipment, "waitlist_queue_depth", None) or 0,
                "entries": result,
                "count": len(result),
                "active_count": active_count,
                "cannot_fulfill_count": cannot_fulfill_count,
            })

        @action(detail=True, methods=["get"], url_path="booking-requesters")
        def booking_requesters(self, request, pk=None):
            """
            Return recipients for equipment email: booking requesters + OIC + Lab In-Charge.
            """
            from iic_booking.users.models.user_group import UserGroupMember

            equipment = self.get_object()
            seen = set()
            recipients = []

            def _add(user_obj, role: str):
                if not user_obj:
                    return
                email = (getattr(user_obj, "email", "") or "").strip()
                if not email or email.lower() in seen:
                    return
                seen.add(email.lower())
                recipients.append(
                    {
                        "user_id": getattr(user_obj, "id", None),
                        "email": email,
                        "name": getattr(user_obj, "name", None) or email,
                        "role": role,
                    }
                )

            # Officer In Charge (managers)
            for em in equipment.equipment_managers.select_related("manager").all():
                _add(getattr(em, "manager", None), "oic")
            # Lab In-Charge (operators)
            for eo in equipment.equipment_operators.select_related("operator").all():
                _add(getattr(eo, "operator", None), "lab")

            link = (
                EquipmentUserGroup.objects.filter(
                    equipment=equipment,
                    purpose=EquipmentUserGroupPurpose.BOOKING_REQUESTERS,
                )
                .select_related("user_group")
                .first()
            )
            group_code = None
            group_name = None
            if link and link.user_group:
                group_code = link.user_group.code
                group_name = link.user_group.name
                members = (
                    UserGroupMember.objects.filter(user_group=link.user_group)
                    .select_related("user")
                    .order_by("user__email")
                )
                for m in members:
                    _add(getattr(m, "user", None), "booking_requester")

            return Response(
                {
                    "equipment_id": equipment.equipment_id,
                    "equipment_code": equipment.code,
                    "equipment_name": getattr(equipment, "name", "") or equipment.code,
                    "group_code": group_code,
                    "group_name": group_name,
                    "recipients": recipients,
                    "count": len(recipients),
                    "oic_count": sum(1 for r in recipients if r.get("role") == "oic"),
                    "lab_count": sum(1 for r in recipients if r.get("role") == "lab"),
                }
            )

        @action(detail=True, methods=["post"], url_path="waitlist-clear")
        def waitlist_clear(self, request, pk=None):
            """Clear the waitlist for this equipment (admin/OIC)."""
            equipment = self.get_object()
            deleted, _ = WaitlistEntry.objects.filter(equipment=equipment).delete()
            return Response({
                "message": f"Waitlist cleared. Removed {deleted} entry(ies).",
                "deleted": deleted,
            }, status=status.HTTP_200_OK)

    class EquipmentCategoryViewSet(ModelViewSet):
        permission_classes = [IsAdminPanelUser]
        queryset = EquipmentCategory.objects.all()
        serializer_class = EquipmentCategorySerializer

    class EquipmentGroupViewSet(ModelViewSet):
        permission_classes = [IsAdminPanelUser]
        queryset = EquipmentGroup.objects.all()
        serializer_class = EquipmentGroupSerializer
        lookup_url_kwarg = "pk"
        lookup_field = "equipment_group_id"

        def get_queryset(self):
            qs = EquipmentGroup.objects.all()
            # Non–Main Admin: only groups that have equipment in their department
            scope_department_id = _request_user_scope_id(self.request)
            if scope_department_id is not None:
                qs = qs.filter(equipment__internal_department_id=scope_department_id).distinct()
            if self.action in ("retrieve", "update", "partial_update"):
                qs = qs.prefetch_related("equipment", "quotas")
            return qs

        def get_serializer_class(self):
            if self.action in ("retrieve", "update", "partial_update"):
                return EquipmentGroupDetailSerializer
            return EquipmentGroupSerializer

        def update(self, request, *args, **kwargs):
            partial = kwargs.get("partial", False)
            instance = self.get_object()
            data = request.data or {}
            for field in ("name", "code", "description"):
                if field in data:
                    setattr(instance, field, data[field])
            instance.save()
            equipment_ids = data.get("equipment_ids")
            if equipment_ids is not None:
                if not isinstance(equipment_ids, list):
                    equipment_ids = []
                valid_ids = []
                for eid in equipment_ids:
                    try:
                        valid_ids.append(int(eid))
                    except (TypeError, ValueError):
                        pass
                Equipment.objects.filter(equipment_group=instance).exclude(equipment_id__in=valid_ids).update(equipment_group=None)
                if valid_ids:
                    Equipment.objects.filter(equipment_id__in=valid_ids).update(equipment_group=instance)
            quotas_data = data.get("quotas")
            if quotas_data is not None and isinstance(quotas_data, list):
                seen_quota_types = set()
                for q in quotas_data:
                    if not isinstance(q, dict):
                        continue
                    quota_type = (q.get("quota_type") or "").strip().upper()
                    if quota_type not in ("WEEKLY", "MONTHLY"):
                        continue
                    seen_quota_types.add(quota_type)
                    obj, _ = EquipmentGroupQuota.objects.get_or_create(
                        equipment_group=instance,
                        quota_type=quota_type,
                        defaults={
                            "internal_individual_quota_minutes": 0,
                            "internal_faculty_quota_minutes": 0,
                            "external_individual_quota_minutes": 0,
                            "external_faculty_quota_minutes": 0,
                            "is_enforced": True,
                        },
                    )
                    if "internal_individual_quota_minutes" in q:
                        obj.internal_individual_quota_minutes = int(q["internal_individual_quota_minutes"]) if q["internal_individual_quota_minutes"] is not None else 0
                    if "internal_faculty_quota_minutes" in q:
                        obj.internal_faculty_quota_minutes = int(q["internal_faculty_quota_minutes"]) if q["internal_faculty_quota_minutes"] is not None else 0
                    if "external_individual_quota_minutes" in q:
                        obj.external_individual_quota_minutes = int(q["external_individual_quota_minutes"]) if q["external_individual_quota_minutes"] is not None else 0
                    if "external_faculty_quota_minutes" in q:
                        obj.external_faculty_quota_minutes = int(q["external_faculty_quota_minutes"]) if q["external_faculty_quota_minutes"] is not None else 0
                    if "is_enforced" in q:
                        obj.is_enforced = bool(q["is_enforced"])
                    obj.save()
                instance.quotas.exclude(quota_type__in=seen_quota_types).delete()
            serializer = EquipmentGroupDetailSerializer(instance, context={"request": request})
            return Response(serializer.data)

    class HolidayViewSet(ModelViewSet):
        permission_classes = [IsAdminPanelUser]
        queryset = Holiday.objects.all().order_by("date")
        serializer_class = HolidaySerializer

        @action(detail=False, methods=["post"], url_path="fetch-gazetted")
        def fetch_gazetted(self, request):
            """Fetch Indian central government gazetted holidays and add new ones."""
            from datetime import date
            import holidays as holidays_lib

            body = getattr(request, "data", None) or {}
            year = body.get("year")
            years = body.get("years")
            if years is not None and not isinstance(years, list):
                years = [years]
            if year is not None and years is None:
                years = [int(year)]
            if not years:
                today = timezone.localdate()
                years = [today.year, today.year + 1]
            try:
                years = [int(y) for y in years]
            except (TypeError, ValueError):
                return Response(
                    {"error": "Invalid year or years. Use integers, e.g. {\"year\": 2025} or {\"years\": [2024, 2025]}."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            if min(years) < 2000 or max(years) > 2030:
                return Response(
                    {"error": "Years must be between 2000 and 2030."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            try:
                in_holidays = holidays_lib.India(years=years)
            except Exception as e:
                return Response(
                    {"error": f"Failed to load holiday data: {e!s}"},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                )
            added = 0
            skipped = 0
            created_holidays = []
            for d, name in sorted(in_holidays.items()):
                if Holiday.objects.filter(date=d).exists():
                    skipped += 1
                    continue
                reason = (name or "").strip()[:255] or str(d)
                holiday = Holiday.objects.create(
                    date=d,
                    reason=reason,
                    is_active=True,
                    color="#fef3c7",
                )
                added += 1
                created_holidays.append(HolidaySerializer(holiday).data)
            return Response(
                {
                    "added": added,
                    "skipped": skipped,
                    "years": years,
                    "holidays": created_holidays,
                },
                status=status.HTTP_200_OK,
            )

        @action(detail=False, methods=["post"], url_path="fetch-from-url")
        def fetch_from_url(self, request):
            """Fetch holidays from a user-defined URL. Expects HTML table with 'Holidays' and 'Date' columns."""
            import re
            from datetime import datetime as dt, date
            from urllib.parse import urlparse
            import requests
            from bs4 import BeautifulSoup

            body = getattr(request, "data", None) or {}
            url = (body.get("url") or "").strip()
            if not url:
                return Response(
                    {"error": "Missing 'url'. Provide a URL to a page containing a table with 'Holidays' and 'Date' columns."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            parsed = urlparse(url)
            if parsed.scheme not in ("http", "https"):
                return Response(
                    {"error": "URL must use http or https."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            try:
                resp = requests.get(
                    url,
                    timeout=15,
                    headers={"User-Agent": "Mozilla/5.0 (compatible; IIC-Booking/1.0)"},
                )
                resp.raise_for_status()
                html = resp.text
            except requests.RequestException as e:
                return Response(
                    {"error": f"Failed to fetch URL: {e!s}"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            soup = BeautifulSoup(html, "html.parser")
            tables = soup.find_all("table")
            default_year = timezone.localdate().year
            year_match = re.search(r"Calendar Year\s*(\d{4})|year\s+(\d{4})|(\d{4})\s*\)", html, re.IGNORECASE)
            if year_match:
                for g in year_match.groups():
                    if g:
                        y = int(g)
                        if 2000 <= y <= 2030:
                            default_year = y
                            break
            date_formats_with_year = [
                "%d/%m/%Y",
                "%d-%m-%Y",
                "%Y-%m-%d",
                "%d.%m.%Y",
                "%B %d, %Y",
                "%d %B %Y",
                "%b %d, %Y",
                "%d %b %Y",
                "%d/%m/%y",
                "%d-%m-%y",
                "%B %d %Y",
                "%b %d %Y",
            ]
            date_formats_month_day = ["%B %d", "%b %d", "%d %B", "%d %b"]
            added = 0
            skipped = 0
            errors = []
            created_holidays = []

            def normalize_header(txt):
                return (txt or "").strip().lower().replace(" ", "").replace("_", "").replace(".", "")

            def parse_date(s, year=default_year):
                s = (s or "").strip()
                if not s:
                    return None
                for fmt in date_formats_with_year:
                    try:
                        return dt.strptime(s, fmt).date()
                    except ValueError:
                        continue
                for fmt in date_formats_month_day:
                    try:
                        parsed = dt.strptime(s, fmt)
                        return date(year, parsed.month, parsed.day)
                    except ValueError:
                        continue
                match = re.search(r"(\d{1,2})[/\-.](\d{1,2})[/\-.](\d{2,4})", s)
                if match:
                    d, m, y = match.groups()
                    y = int(y)
                    if y < 100:
                        y += 2000 if y < 50 else 1900
                    try:
                        return date(int(y), int(m), int(d))
                    except ValueError:
                        pass
                return None

            holidays_col = normalize_header("Holidays")
            date_col = normalize_header("Date")
            found_any = False

            for table in tables:
                rows = table.find_all("tr")
                if len(rows) < 2:
                    continue
                header_cells = rows[0].find_all(["th", "td"])
                if not header_cells:
                    continue
                indices = {}
                for i, cell in enumerate(header_cells):
                    key = normalize_header(cell.get_text())
                    if key:
                        indices[key] = i
                idx_holidays = indices.get(holidays_col)
                idx_date = indices.get(date_col)
                if idx_holidays is None or idx_date is None:
                    continue
                found_any = True
                for row in rows[1:]:
                    cells = row.find_all(["td", "th"])
                    if max(idx_holidays, idx_date) >= len(cells):
                        continue
                    reason_raw = (cells[idx_holidays].get_text() or "").strip()
                    if reason_raw.endswith("*"):
                        reason_raw = reason_raw[:-1].strip()
                    reason_raw = reason_raw[:255]
                    date_raw = (cells[idx_date].get_text() or "").strip()
                    if not reason_raw and not date_raw:
                        continue
                    parsed_date = parse_date(date_raw, default_year)
                    if not parsed_date:
                        errors.append(f"Could not parse date '{date_raw}' for '{reason_raw}'")
                        continue
                    if Holiday.objects.filter(date=parsed_date).exists():
                        skipped += 1
                        continue
                    reason = reason_raw or str(parsed_date)
                    holiday = Holiday.objects.create(
                        date=parsed_date,
                        reason=reason,
                        is_active=True,
                        color="#fef3c7",
                    )
                    added += 1
                    created_holidays.append(HolidaySerializer(holiday).data)

            if not found_any:
                return Response(
                    {
                        "error": "No table with both 'Holidays' and 'Date' columns found on this page.",
                        "added": 0,
                        "skipped": 0,
                        "errors": errors,
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )
            return Response(
                {
                    "added": added,
                    "skipped": skipped,
                    "errors": errors,
                    "holidays": created_holidays,
                    "year_used": default_year,
                },
                status=status.HTTP_200_OK,
            )

    class MenuItemViewSet(ModelViewSet):
        permission_classes = [IsAdminPanelUser]
        queryset = MenuItem.objects.select_related("page").order_by("priority", "id")
        serializer_class = MenuItemListSerializer

        def get_serializer_class(self):
            if self.action == "list":
                return MenuItemListSerializer
            return MenuItemSerializer

    class CmsPageViewSet(ModelViewSet):
        permission_classes = [IsAdminPanelUser]
        queryset = CmsPage.objects.all().order_by("title")
        serializer_class = CmsPageSerializer

        @action(detail=False, methods=["post"], url_path="upload-image")
        def upload_image(self, request):
            """Upload an image for use in CMS page blocks. Returns { url: absolute_url }."""
            file = request.FILES.get("image") or request.FILES.get("file")
            if not file:
                return Response(
                    {"error": "No image file provided. Use form field 'image' or 'file'."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            # Restrict to image types
            allowed = ("image/jpeg", "image/png", "image/gif", "image/webp")
            if getattr(file, "content_type", "") not in allowed:
                return Response(
                    {"error": "Invalid file type. Use JPEG, PNG, GIF, or WebP."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            ext = os.path.splitext(getattr(file, "name", ""))[1] or ".jpg"
            filename = f"cms/page_images/{datetime.now().strftime('%Y/%m')}/page_{timestamp}{ext}"
            try:
                file.seek(0)
                saved_path = default_storage.save(filename, ContentFile(file.read()))
                url = default_storage.url(saved_path)
                if url.startswith("/"):
                    url = request.build_absolute_uri(url)
                return Response({"url": url})
            except Exception as e:
                return Response(
                    {"error": f"Upload failed: {str(e)}"},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                )

        @action(detail=False, methods=["post"], url_path="upload-document")
        def upload_document(self, request):
            """Upload a document (e.g. resume PDF) for use in CMS page blocks. Returns { url: absolute_url }."""
            file = request.FILES.get("document") or request.FILES.get("file")
            if not file:
                return Response(
                    {"error": "No file provided. Use form field 'document' or 'file'."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            allowed = (
                "application/pdf",
                "application/msword",
                "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            )
            if getattr(file, "content_type", "") not in allowed:
                return Response(
                    {"error": "Invalid file type. Use PDF or Word (DOC/DOCX)."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            ext = os.path.splitext(getattr(file, "name", ""))[1] or ".pdf"
            filename = f"cms/page_documents/{datetime.now().strftime('%Y/%m')}/doc_{timestamp}{ext}"
            try:
                file.seek(0)
                saved_path = default_storage.save(filename, ContentFile(file.read()))
                url = default_storage.url(saved_path)
                if url.startswith("/"):
                    url = request.build_absolute_uri(url)
                return Response({"url": url})
            except Exception as e:
                return Response(
                    {"error": f"Upload failed: {str(e)}"},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                )

    class HomePageContentViewSet(ModelViewSet):
        permission_classes = [IsAdminPanelUser]
        queryset = HomePageContent.objects.all().order_by("key")
        serializer_class = HomePageContentSerializer

    class HeroSlideViewSet(ModelViewSet):
        permission_classes = [IsAdminPanelUser]
        queryset = HeroSlide.objects.all().order_by("order", "id")
        serializer_class = HeroSlideSerializer
        parser_classes = [JSONParser, FormParser, MultiPartParser]

    # Communication (admin-only: templates + logs)
    from iic_booking.communication.models import CommunicationTemplate, CommunicationLog
    from django.db.models import Q as DQ

    class CommunicationTemplateSerializer(serializers.ModelSerializer):
        created_by_email = serializers.EmailField(source="created_by.email", read_only=True, allow_null=True)
        updated_by_email = serializers.EmailField(source="updated_by.email", read_only=True, allow_null=True)
        communication_type_display = serializers.CharField(source="get_communication_type_display", read_only=True)

        class Meta:
            model = CommunicationTemplate
            fields = [
                "id", "name", "code", "communication_type", "communication_type_display",
                "description", "subject", "body_text", "body_html", "sms_body", "push_data",
                "variable_help", "is_active", "created_at", "updated_at",
                "created_by", "updated_by", "created_by_email", "updated_by_email",
            ]
            read_only_fields = ["id", "created_at", "updated_at", "created_by_email", "updated_by_email", "communication_type_display"]

    class CommunicationTemplateViewSet(ModelViewSet):
        permission_classes = [IsAdminOrDeptCommunicationAdmin]
        serializer_class = CommunicationTemplateSerializer
        queryset = CommunicationTemplate.objects.all().select_related("created_by", "updated_by").order_by("communication_type", "name")

        def get_queryset(self):
            qs = super().get_queryset()
            search = self.request.query_params.get("search", "").strip()
            if search:
                qs = qs.filter(
                    DQ(name__icontains=search)
                    | DQ(code__icontains=search)
                    | DQ(subject__icontains=search)
                    | DQ(description__icontains=search)
                )
            comm_type = self.request.query_params.get("communication_type", "").strip()
            if comm_type:
                qs = qs.filter(communication_type=comm_type)
            is_active = self.request.query_params.get("is_active")
            if is_active is not None:
                if str(is_active).lower() in ("true", "1", "yes"):
                    qs = qs.filter(is_active=True)
                elif str(is_active).lower() in ("false", "0", "no"):
                    qs = qs.filter(is_active=False)
            return qs

        def perform_create(self, serializer):
            serializer.save(created_by=self.request.user, updated_by=self.request.user)

        def perform_update(self, serializer):
            serializer.save(updated_by=self.request.user)

    class CommunicationLogSerializer(serializers.ModelSerializer):
        recipient_email = serializers.EmailField(source="recipient.email", read_only=True)
        recipient_name = serializers.CharField(source="recipient.name", read_only=True, allow_null=True)
        communication_type_display = serializers.CharField(source="get_communication_type_display", read_only=True)
        status_display = serializers.CharField(source="get_status_display", read_only=True)
        template_code = serializers.CharField(source="template.code", read_only=True, allow_null=True)

        class Meta:
            model = CommunicationLog
            fields = [
                "id", "communication_type", "communication_type_display", "recipient", "recipient_email", "recipient_name",
                "template", "template_code", "subject", "message", "status", "status_display",
                "sent_at", "delivered_at", "read_at", "error_message", "provider_message_id",
                "metadata", "created_by", "created_at", "updated_at",
            ]
            read_only_fields = fields

    class CommunicationLogViewSet(ModelViewSet):
        permission_classes = [IsAdminOrDeptCommunicationAdmin]
        serializer_class = CommunicationLogSerializer
        queryset = CommunicationLog.objects.all().select_related("recipient", "template", "created_by").order_by("-created_at")
        http_method_names = ["get", "head", "options"]

        def get_queryset(self):
            from datetime import datetime
            qs = super().get_queryset()
            search = self.request.query_params.get("search", "").strip()
            if search:
                qs = qs.filter(
                    DQ(recipient__email__icontains=search)
                    | DQ(recipient__name__icontains=search)
                    | DQ(subject__icontains=search)
                    | DQ(message__icontains=search)
                    | DQ(error_message__icontains=search)
                    | DQ(provider_message_id__icontains=search)
                )
            comm_type = self.request.query_params.get("communication_type", "").strip()
            if comm_type:
                qs = qs.filter(communication_type=comm_type)
            status_filter = self.request.query_params.get("status", "").strip()
            if status_filter:
                qs = qs.filter(status=status_filter)
            date_from = self.request.query_params.get("date_from", "").strip()
            if date_from:
                try:
                    parsed = datetime.strptime(date_from, "%Y-%m-%d").date()
                    start_dt = timezone.make_aware(datetime.combine(parsed, datetime.min.time()))
                    qs = qs.filter(created_at__gte=start_dt)
                except ValueError:
                    pass
            date_to = self.request.query_params.get("date_to", "").strip()
            if date_to:
                try:
                    parsed = datetime.strptime(date_to, "%Y-%m-%d").date()
                    end_dt = timezone.make_aware(datetime.combine(parsed + timedelta(days=1), datetime.min.time()))
                    qs = qs.filter(created_at__lt=end_dt)
                except ValueError:
                    pass
            return qs

    # Notices (admin-only, same as Django admin /admin/communication/notice/)
    from iic_booking.communication.models import Notice

    class NoticeAdminSerializer(serializers.ModelSerializer):
        created_by_name = serializers.CharField(source="created_by.name", read_only=True, allow_null=True)
        created_by_email = serializers.EmailField(source="created_by.email", read_only=True, allow_null=True)
        notice_type_display = serializers.CharField(source="get_notice_type_display", read_only=True)

        class Meta:
            model = Notice
            fields = [
                "notice_id", "title", "description", "content", "notice_type", "notice_type_display",
                "is_active", "priority", "expiry_date", "created_by", "created_by_name", "created_by_email",
                "created_at", "updated_at",
            ]
            read_only_fields = ["notice_id", "created_at", "updated_at", "created_by_name", "created_by_email", "notice_type_display"]

    class NoticeViewSet(ModelViewSet):
        permission_classes = [IsAdminOrDeptCommunicationAdmin]
        serializer_class = NoticeAdminSerializer
        queryset = Notice.objects.all().select_related("created_by").order_by("-priority", "-created_at")

        def get_queryset(self):
            qs = super().get_queryset()
            search = self.request.query_params.get("search", "").strip()
            if search:
                qs = qs.filter(
                    DQ(title__icontains=search)
                    | DQ(description__icontains=search)
                    | DQ(content__icontains=search)
                )
            notice_type = self.request.query_params.get("notice_type", "").strip()
            if notice_type:
                qs = qs.filter(notice_type=notice_type)
            is_active = self.request.query_params.get("is_active")
            if is_active is not None:
                if str(is_active).lower() in ("true", "1", "yes"):
                    qs = qs.filter(is_active=True)
                elif str(is_active).lower() in ("false", "0", "no"):
                    qs = qs.filter(is_active=False)
            return qs

        def perform_create(self, serializer):
            serializer.save(created_by=self.request.user)

    router = DefaultRouter()
    router.register(r"departments", DepartmentViewSet, basename="admin-department")
    router.register(r"organization-requests", OrganizationRequestViewSet, basename="admin-organization-request")
    router.register(r"projects", ProjectViewSet, basename="admin-project")
    router.register(r"users", UserAdminViewSet, basename="admin-user")
    router.register(r"permission-definitions", PermissionDefinitionViewSet, basename="admin-permission-definition")
    router.register(r"dept-admin-grants", DeptAdminPermissionGrantViewSet, basename="admin-dept-admin-grant")
    from config.admin_panel_access_api import AdminPanelRoleConfigViewSet

    router.register(
        r"admin-panel-role-configs",
        AdminPanelRoleConfigViewSet,
        basename="admin-panel-role-config",
    )
    router.register(r"staff-permission-grants", StaffPermissionGrantViewSet, basename="admin-staff-permission-grant")
    router.register(r"user-groups", UserGroupViewSet, basename="admin-usergroup")
    router.register(r"user-group-members", UserGroupMemberViewSet, basename="admin-usergroupmember")
    router.register(r"user-documents", UserDocumentViewSet, basename="admin-userdocument")
    router.register(r"wallets", WalletViewSet, basename="admin-wallet")
    router.register(r"sub-wallets", SubWalletViewSet, basename="admin-subwallet")
    router.register(r"sub-wallet-transactions", SubWalletTransactionViewSet, basename="admin-subwallettransaction")
    router.register(r"wallet-razorpay-orders", WalletRazorpayOrderViewSet, basename="admin-walletrazorpayorder")
    router.register(r"wallet-recharge-requests", WalletRechargeRequestViewSet, basename="admin-walletrechargerequest")
    router.register(r"bookings", BookingViewSet, basename="admin-booking")
    router.register(r"repeat-sample-requests", RepeatSampleRequestViewSet, basename="admin-repeat-sample-request")
    router.register(r"daily-slots", DailySlotViewSet, basename="admin-dailyslot")
    router.register(r"equipment", EquipmentViewSet, basename="admin-equipment")
    router.register(r"equipment-categories", EquipmentCategoryViewSet, basename="admin-equipmentcategory")
    router.register(r"equipment-groups", EquipmentGroupViewSet, basename="admin-equipmentgroup")
    router.register(r"holidays", HolidayViewSet, basename="admin-holiday")
    router.register(r"cms-menu", MenuItemViewSet, basename="admin-cms-menu")
    router.register(r"cms-pages", CmsPageViewSet, basename="admin-cms-pages")
    router.register(r"cms-home", HomePageContentViewSet, basename="admin-cms-home")
    router.register(r"cms-hero-slides", HeroSlideViewSet, basename="admin-cms-hero-slides")
    router.register(r"communication-templates", CommunicationTemplateViewSet, basename="admin-communication-template")
    router.register(r"communication-logs", CommunicationLogViewSet, basename="admin-communication-log")
    router.register(r"notices", NoticeViewSet, basename="admin-notice")

    # Calendar colors for weekly window (admin-only: get and update)
    class CalendarColorViewSet(ViewSet):
        permission_classes = [IsAdminUser]

        def list(self, request):
            return Response(get_calendar_colors())

        @action(detail=False, methods=["patch"], url_path="update")
        def update_colors(self, request):
            data = request.data or {}
            slot_colors = data.get("slot_colors") or {}
            holiday_default = data.get("holiday_default")
            saturday_color = data.get("saturday_color")
            sunday_color = data.get("sunday_color")
            slot_keys = {
                "AVAILABLE",
                "BOOKED",
                "COMPLETED",
                "HOLD",
                "BLOCKED",
                "UNDER_MAINTENANCE",
                "OPERATOR_ABSENT",
                "BOOKING_NOT_UTILIZED",
                "RESERVED_FOR_EXTERNAL",
                "HOME_DEPARTMENT_ONLY",
                "NON_HOME_RESERVED",
                "NOT_AVAILABLE",
            }
            updated = []
            for key, value in slot_colors.items():
                if key in slot_keys and value and isinstance(value, str) and value.strip().startswith("#"):
                    obj, _ = CalendarColorSetting.objects.get_or_create(key=key, defaults={"value": value.strip()})
                    obj.value = value.strip()
                    obj.save()
                    updated.append(key)
            if holiday_default and isinstance(holiday_default, str) and holiday_default.strip().startswith("#"):
                obj, _ = CalendarColorSetting.objects.get_or_create(key="HOLIDAY_DEFAULT", defaults={"value": holiday_default.strip()})
                obj.value = holiday_default.strip()
                obj.save()
                updated.append("HOLIDAY_DEFAULT")
            if saturday_color and isinstance(saturday_color, str) and saturday_color.strip().startswith("#"):
                obj, _ = CalendarColorSetting.objects.get_or_create(key="SATURDAY", defaults={"value": saturday_color.strip()})
                obj.value = saturday_color.strip()
                obj.save()
                updated.append("SATURDAY")
            if sunday_color and isinstance(sunday_color, str) and sunday_color.strip().startswith("#"):
                obj, _ = CalendarColorSetting.objects.get_or_create(key="SUNDAY", defaults={"value": sunday_color.strip()})
                obj.value = sunday_color.strip()
                obj.save()
                updated.append("SUNDAY")
            external_gst = data.get("external_gst_percent")
            if external_gst is not None:
                try:
                    gst_val = float(external_gst)
                    if 0 <= gst_val <= 100:
                        obj, _ = BookingChargeSetting.objects.get_or_create(
                            key="EXTERNAL_GST_PERCENT", defaults={"value": str(int(round(gst_val)))}
                        )
                        obj.value = str(int(round(gst_val)))
                        obj.save()
                        updated.append("EXTERNAL_GST_PERCENT")
                except (TypeError, ValueError):
                    pass
            return Response(get_calendar_colors())

    router.register(r"calendar-colors", CalendarColorViewSet, basename="admin-calendar-colors")

    # Internal user slot window (common for all equipment) – admin only
    class InternalUserSlotWindowViewSet(ViewSet):
        permission_classes = [IsAdminUser]

        def list(self, request):
            """GET: return the single internal slot window setting (singleton)."""
            obj, _ = InternalUserSlotWindowSetting.objects.get_or_create(
                defaults={"reference_weekday": None, "reference_time": None}
            )
            return Response({
                "reference_weekday": obj.reference_weekday,
                "reference_time": obj.reference_time.strftime("%H:%M") if obj.reference_time else None,
            })

        @action(detail=False, methods=["patch", "put"], url_path="update")
        def update_setting(self, request):
            """PATCH/PUT: update reference_weekday and reference_time."""
            obj, _ = InternalUserSlotWindowSetting.objects.get_or_create(
                defaults={"reference_weekday": None, "reference_time": None}
            )
            data = request.data or {}
            if "reference_weekday" in data:
                v = data["reference_weekday"]
                if v is None or (isinstance(v, str) and not v.strip()):
                    obj.reference_weekday = None
                else:
                    try:
                        n = int(v)
                        obj.reference_weekday = n if 0 <= n <= 6 else None
                    except (TypeError, ValueError):
                        pass
            if "reference_time" in data:
                v = data["reference_time"]
                if v is None or (isinstance(v, str) and not v.strip()):
                    obj.reference_time = None
                else:
                    from datetime import datetime
                    s = (v if isinstance(v, str) else str(v)).strip()[:5]  # HH:mm
                    try:
                        obj.reference_time = datetime.strptime(s, "%H:%M").time()
                    except ValueError:
                        pass
            obj.save()
            return Response({
                "reference_weekday": obj.reference_weekday,
                "reference_time": obj.reference_time.strftime("%H:%M") if obj.reference_time else None,
            })

    router.register(r"internal-slot-window", InternalUserSlotWindowViewSet, basename="admin-internal-slot-window")

    # Equipment utilization reports (data + PDF/Excel download)
    from iic_booking.equipment.reports import get_equipment_report_data
    from iic_booking.equipment.report_exports import build_report_pdf, build_report_excel
    from django.http import HttpResponse

    class EquipmentReportViewSet(ViewSet):
        permission_classes = [IsAdminPanelUserOrReportsStaff]

        def initial(self, request, *args, **kwargs):
            super().initial(request, *args, **kwargs)
            _require_reports_view(request)

        def _resolve_equipment_ids(self, request):
            """Return equipment_ids for report: restrict to user's equipment for manager/operator."""
            from iic_booking.equipment.reports import get_equipment_ids_managed_by_oic
            from iic_booking.equipment.api_views import _get_equipment_ids_for_log_access
            equipment_ids = request.query_params.getlist("equipment_id")
            if equipment_ids:
                try:
                    equipment_ids = [int(x) for x in equipment_ids if x]
                except ValueError:
                    equipment_ids = None
            else:
                equipment_ids = None
            if getattr(request.user, "user_type", None) == UserType.MANAGER:
                allowed = get_equipment_ids_managed_by_oic(request.user.id)
                if not allowed:
                    equipment_ids = []
                elif equipment_ids is None:
                    equipment_ids = allowed
                else:
                    equipment_ids = [x for x in equipment_ids if x in allowed]
            elif getattr(request.user, "user_type", None) == UserType.OPERATOR:
                allowed = _get_equipment_ids_for_log_access(request.user) or []
                if not allowed:
                    equipment_ids = []
                elif equipment_ids is None:
                    equipment_ids = allowed
                else:
                    equipment_ids = [x for x in equipment_ids if x in allowed]
            return equipment_ids

        def list(self, request):
            """GET ?date_from=YYYY-MM-DD&date_to=YYYY-MM-DD&equipment_id=1&equipment_id=2 - report data."""
            date_from = request.query_params.get("date_from") or None
            date_to = request.query_params.get("date_to") or None
            equipment_ids = self._resolve_equipment_ids(request)
            data = get_equipment_report_data(
                date_from=date_from,
                date_to=date_to,
                equipment_ids=equipment_ids,
            )
            return Response(data)

        @action(detail=False, methods=["get"], url_path="download-pdf")
        def download_pdf(self, request):
            """GET ?date_from=...&date_to=...&equipment_id=1 - download report as PDF."""
            date_from = request.query_params.get("date_from") or None
            date_to = request.query_params.get("date_to") or None
            equipment_ids = self._resolve_equipment_ids(request)
            pdf_bytes = build_report_pdf(
                date_from=date_from,
                date_to=date_to,
                equipment_ids=equipment_ids,
            )
            filename = f"equipment-report-{timezone.localdate().isoformat()}.pdf"
            resp = HttpResponse(pdf_bytes, content_type="application/pdf")
            resp["Content-Disposition"] = f'attachment; filename="{filename}"'
            return resp

        @action(detail=False, methods=["get"], url_path="download-excel")
        def download_excel(self, request):
            """GET ?date_from=...&date_to=...&equipment_id=1 - download report as Excel."""
            date_from = request.query_params.get("date_from") or None
            date_to = request.query_params.get("date_to") or None
            equipment_ids = self._resolve_equipment_ids(request)
            xlsx_bytes = build_report_excel(
                date_from=date_from,
                date_to=date_to,
                equipment_ids=equipment_ids,
            )
            filename = f"equipment-report-{timezone.localdate().isoformat()}.xlsx"
            resp = HttpResponse(xlsx_bytes, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            resp["Content-Disposition"] = f'attachment; filename="{filename}"'
            return resp

    router.register(r"equipment-reports", EquipmentReportViewSet, basename="admin-equipment-reports")

    from config.admin_extra_settings_api import register_extra_admin_routes

    register_extra_admin_routes(router)

    return router
