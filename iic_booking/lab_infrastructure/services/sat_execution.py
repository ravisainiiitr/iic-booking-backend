"""Lab SAT Execution Mode — wizard order, failure capture, readiness, reports."""

from __future__ import annotations

import csv
import io
import json
from typing import Any

from django.db.models import Count, Q
from django.http import HttpResponse
from django.utils import timezone

from iic_booking.lab_infrastructure.models import (
    LabAlert,
    SatDefect,
    SatEvidence,
    SatTestCase,
    SatTestResult,
    SatTestRun,
)

STAGE_LABELS = {
    1: "Stage 1 — Deployment & Commissioning",
    2: "Stage 2 — DSA / Equipment PC / Sync",
    3: "Stage 3 — Booking & Remote Analysis",
    4: "Stage 4 — Fleet / Diagnostics / Reporting",
    5: "Stage 5 — Performance / Security / Final",
}

# Explicit Lab SAT execution order (user Stage 1→5).
EXECUTION_PLAN: list[tuple[str, int, int]] = [
    # Stage 1
    ("SAT-DEP-001", 1, 10),
    ("SAT-DEP-002", 1, 20),
    ("SAT-DEP-003", 1, 30),
    ("SAT-COM-001", 1, 40),
    ("SAT-COM-002", 1, 50),
    ("SAT-COM-003", 1, 60),
    # Stage 2
    ("SAT-COM-004", 2, 70),
    ("SAT-COM-005", 2, 80),
    ("SAT-COM-006", 2, 90),
    ("SAT-DSA-001", 2, 100),
    ("SAT-DSA-002", 2, 110),
    ("SAT-DSA-003", 2, 120),
    ("SAT-DSA-004", 2, 130),
    ("SAT-DSA-005", 2, 140),
    ("SAT-BKG-006", 2, 150),
    ("SAT-FAIL-001", 2, 160),
    ("SAT-FAIL-002", 2, 170),
    ("SAT-FAIL-004", 2, 180),
    ("SAT-FAIL-005", 2, 190),
    # Stage 3
    ("SAT-BKG-001", 3, 200),
    ("SAT-BKG-002", 3, 210),
    ("SAT-BKG-003", 3, 220),
    ("SAT-BKG-004", 3, 230),
    ("SAT-BKG-005", 3, 240),
    ("SAT-RA-001", 3, 250),
    ("SAT-RA-002", 3, 260),
    ("SAT-RA-003", 3, 270),
    ("SAT-RA-004", 3, 280),
    ("SAT-RA-005", 3, 290),
    ("SAT-RA-006", 3, 300),
    ("SAT-FAIL-003", 3, 310),
    # Stage 4
    ("SAT-DEP-004", 4, 320),
    ("SAT-DEP-005", 4, 330),
    ("SAT-DEP-006", 4, 340),
    ("SAT-FLT-001", 4, 350),
    ("SAT-FLT-002", 4, 360),
    ("SAT-FLT-003", 4, 370),
    ("SAT-FLT-004", 4, 380),
    ("SAT-FLT-005", 4, 390),
    ("SAT-FE-001", 4, 400),
    ("SAT-FE-002", 4, 410),
    # Stage 5
    ("SAT-PERF-001", 5, 420),
    ("SAT-PERF-002", 5, 430),
    ("SAT-SEC-001", 5, 440),
    ("SAT-SEC-002", 5, 450),
    ("SAT-SEC-003", 5, 460),
    ("SAT-SEC-004", 5, 470),
    ("SAT-API-001", 5, 480),
    ("SAT-API-002", 5, 490),
    ("SAT-DB-001", 5, 500),
    ("SAT-DB-002", 5, 510),
]

READINESS_DOMAINS = {
    "Deployment": ["Deployment", "Deployment Center", "Commissioning"],
    "Synchronization": ["Synchronization", "DSA", "Configuration Push", "Heartbeat"],
    "Remote Analysis": ["Remote Analysis", "Booking Workflow"],
    "Security": ["Security", "Role Based Access"],
    "Diagnostics": ["Diagnostics", "Repair", "Failure Recovery"],
    "Performance": ["Performance"],
    "Fleet": ["Fleet Dashboard", "Reporting", "Notifications", "Software Inventory"],
    "Platform": ["API", "Database", "Frontend"],
}


def apply_execution_plan() -> None:
    plan = {tid: (stage, order) for tid, stage, order in EXECUTION_PLAN}
    for tc in SatTestCase.objects.filter(is_active=True):
        if tc.test_id in plan:
            stage, order = plan[tc.test_id]
            if tc.stage != stage or tc.execution_order != order:
                tc.stage = stage
                tc.execution_order = order
                tc.save(update_fields=["stage", "execution_order", "updated_at"])


def run_progress(run: SatTestRun) -> dict[str, Any]:
    qs = SatTestResult.objects.filter(run=run)
    agg = qs.aggregate(
        total=Count("id"),
        passed=Count("id", filter=Q(status=SatTestResult.Status.PASSED)),
        failed=Count("id", filter=Q(status=SatTestResult.Status.FAILED)),
        blocked=Count("id", filter=Q(status=SatTestResult.Status.BLOCKED)),
        skipped=Count("id", filter=Q(status=SatTestResult.Status.SKIPPED)),
        not_run=Count("id", filter=Q(status=SatTestResult.Status.NOT_RUN)),
    )
    total = agg["total"] or 0
    done = (agg["passed"] or 0) + (agg["failed"] or 0) + (agg["blocked"] or 0) + (agg["skipped"] or 0)
    completion = round(100.0 * done / total, 1) if total else 0.0
    return {
        "run_id": str(run.id),
        "run_name": run.name,
        "status": run.status,
        "recommendation": run.recommendation,
        "total": total,
        "passed": agg["passed"] or 0,
        "failed": agg["failed"] or 0,
        "blocked": agg["blocked"] or 0,
        "skipped": agg["skipped"] or 0,
        "remaining": agg["not_run"] or 0,
        "completion_pct": completion,
        "started_at": run.started_at.isoformat() if run.started_at else None,
        "finished_at": run.finished_at.isoformat() if run.finished_at else None,
        "lab_context": run.lab_context or {},
    }


def stage_summary(run: SatTestRun) -> list[dict[str, Any]]:
    out = []
    for stage in range(1, 6):
        qs = SatTestResult.objects.filter(run=run, test_case__stage=stage)
        total = qs.count()
        if not total:
            continue
        agg = qs.aggregate(
            passed=Count("id", filter=Q(status=SatTestResult.Status.PASSED)),
            failed=Count("id", filter=Q(status=SatTestResult.Status.FAILED)),
            blocked=Count("id", filter=Q(status=SatTestResult.Status.BLOCKED)),
            skipped=Count("id", filter=Q(status=SatTestResult.Status.SKIPPED)),
            not_run=Count("id", filter=Q(status=SatTestResult.Status.NOT_RUN)),
        )
        done = (agg["passed"] or 0) + (agg["failed"] or 0) + (agg["blocked"] or 0) + (agg["skipped"] or 0)
        out.append(
            {
                "stage": stage,
                "label": STAGE_LABELS.get(stage, f"Stage {stage}"),
                "total": total,
                "passed": agg["passed"] or 0,
                "failed": agg["failed"] or 0,
                "blocked": agg["blocked"] or 0,
                "skipped": agg["skipped"] or 0,
                "remaining": agg["not_run"] or 0,
                "completion_pct": round(100.0 * done / total, 1) if total else 0.0,
            }
        )
    return out


def serialize_result(result: SatTestResult, *, include_evidence: bool = True) -> dict[str, Any]:
    tc = result.test_case
    payload: dict[str, Any] = {
        "id": str(result.id),
        "run_id": str(result.run_id),
        "test_id": tc.test_id,
        "module": tc.module,
        "feature": tc.feature,
        "severity": tc.severity,
        "stage": tc.stage,
        "stage_label": STAGE_LABELS.get(tc.stage, ""),
        "execution_order": tc.execution_order,
        "preconditions": tc.preconditions,
        "steps": tc.steps,
        "expected_result": tc.expected_result,
        "actual_result": result.actual_result,
        "remarks": result.remarks,
        "administrator_notes": result.administrator_notes,
        "log_url": result.log_url,
        "status": result.status,
        "failure_snapshot": result.failure_snapshot or {},
        "executed_at": result.executed_at.isoformat() if result.executed_at else None,
    }
    if include_evidence:
        payload["evidence_files"] = [
            {
                "id": str(e.id),
                "kind": e.kind,
                "title": e.title,
                "original_name": e.original_name,
                "url": e.file.url if e.file else "",
                "created_at": e.created_at.isoformat() if e.created_at else None,
            }
            for e in result.evidence_files.all()[:50]
        ]
        payload["defects"] = [
            {
                "id": str(d.id),
                "kind": d.kind,
                "severity": d.severity,
                "status": d.status,
                "title": d.title,
            }
            for d in result.defects.all()[:20]
        ]
    return payload


def current_wizard_step(run: SatTestRun) -> dict[str, Any] | None:
    """Return the next not_run result in execution order, or current_result if set and unfinished."""
    if run.current_result_id:
        cur = (
            SatTestResult.objects.select_related("test_case")
            .prefetch_related("evidence_files", "defects")
            .filter(pk=run.current_result_id)
            .first()
        )
        if cur and cur.status == SatTestResult.Status.NOT_RUN:
            return serialize_result(cur)

    nxt = (
        SatTestResult.objects.select_related("test_case")
        .prefetch_related("evidence_files", "defects")
        .filter(run=run, status=SatTestResult.Status.NOT_RUN)
        .order_by("test_case__stage", "test_case__execution_order", "test_case__test_id")
        .first()
    )
    if not nxt:
        return None
    run.current_result = nxt
    run.save(update_fields=["current_result"])
    return serialize_result(nxt)


def capture_failure_snapshot(*, notes: str = "", lab_context: dict | None = None) -> dict[str, Any]:
    """Best-effort live environment capture for failure analysis."""
    snapshot: dict[str, Any] = {
        "timestamp": timezone.now().isoformat(),
        "administrator_notes": notes or "",
        "lab_context": lab_context or {},
        "alerts": [],
        "fleet_summary": {},
        "nodes": [],
    }
    try:
        alerts = LabAlert.objects.filter(status=LabAlert.Status.OPEN).order_by("-created_at")[:20]
        snapshot["alerts"] = [
            {
                "code": a.code,
                "severity": a.severity,
                "title": a.title,
                "node_id": a.node_id,
                "created_at": a.created_at.isoformat() if a.created_at else None,
            }
            for a in alerts
        ]
    except Exception as exc:  # noqa: BLE001
        snapshot["alerts_error"] = str(exc)

    try:
        from iic_booking.lab_infrastructure.services.fleet import build_infrastructure_tree

        tree = build_infrastructure_tree(page=1, page_size=200)
        departments = tree.get("departments") or tree.get("tree") or []
        nodes_flat: list[dict] = []

        def walk(node: dict, trail: list[str]) -> None:
            kind = node.get("kind") or node.get("type") or ""
            name = node.get("computer_name") or node.get("name") or node.get("id") or ""
            status = node.get("status") or ""
            path = trail + [str(name)]
            if kind in {"dsa", "equipment_pc", "raa", "analysis_pc"} or node.get("id"):
                nodes_flat.append(
                    {
                        "id": node.get("id"),
                        "kind": kind,
                        "name": name,
                        "status": status,
                        "path": " / ".join(path),
                        "ip_address": node.get("ip_address"),
                        "agent_version": node.get("agent_version"),
                        "configuration_version": node.get("configuration_version"),
                        "health_score": node.get("health_score"),
                        "cpu": node.get("cpu"),
                        "memory": node.get("memory"),
                        "disk": node.get("disk"),
                        "last_heartbeat": node.get("last_heartbeat"),
                        "equipment": node.get("equipment"),
                        "department": node.get("department"),
                    }
                )
            for child in node.get("children") or []:
                if isinstance(child, dict):
                    walk(child, path)

        if isinstance(departments, list):
            for dept in departments:
                if isinstance(dept, dict):
                    walk(dept, [str(dept.get("name") or dept.get("department") or "Dept")])
        offline = sum(1 for n in nodes_flat if str(n.get("status") or "").lower() in {"offline", "error"})
        snapshot["fleet_summary"] = {
            "node_count": len(nodes_flat),
            "offline_or_error": offline,
        }
        snapshot["nodes"] = nodes_flat[:100]
    except Exception as exc:  # noqa: BLE001
        snapshot["fleet_error"] = str(exc)

    return snapshot


def record_result(
    result: SatTestResult,
    *,
    status: str,
    actual_result: str = "",
    remarks: str = "",
    administrator_notes: str = "",
    log_url: str = "",
    advance: bool = True,
) -> SatTestResult:
    result.status = status
    result.executed_at = timezone.now()
    if actual_result:
        result.actual_result = actual_result[:8000]
    if remarks:
        result.remarks = remarks[:4000]
    if administrator_notes:
        result.administrator_notes = administrator_notes[:8000]
    if log_url:
        result.log_url = log_url[:500]
    if status in {SatTestResult.Status.FAILED, SatTestResult.Status.BLOCKED}:
        result.failure_snapshot = capture_failure_snapshot(
            notes=administrator_notes or remarks,
            lab_context=result.run.lab_context if result.run_id else {},
        )
    result.save()

    run = result.run
    if advance:
        nxt = (
            SatTestResult.objects.filter(run=run, status=SatTestResult.Status.NOT_RUN)
            .order_by("test_case__stage", "test_case__execution_order", "test_case__test_id")
            .first()
        )
        run.current_result = nxt
        if nxt is None and run.status == SatTestRun.Status.RUNNING:
            # All cases decided — freeze readiness
            readiness = compute_readiness(run)
            run.readiness_snapshot = readiness
            run.recommendation = readiness.get("recommendation", SatTestRun.Recommendation.PENDING)
            run.status = SatTestRun.Status.COMPLETED
            run.finished_at = timezone.now()
            run.save(
                update_fields=[
                    "current_result",
                    "readiness_snapshot",
                    "recommendation",
                    "status",
                    "finished_at",
                ]
            )
        else:
            run.save(update_fields=["current_result"])
    return result


def domain_score(run: SatTestRun, modules: list[str]) -> dict[str, Any]:
    qs = SatTestResult.objects.filter(run=run, test_case__module__in=modules)
    total = qs.count()
    if not total:
        return {"score": None, "total": 0, "passed": 0, "failed": 0, "remaining": 0}
    passed = qs.filter(status=SatTestResult.Status.PASSED).count()
    failed = qs.filter(status__in=[SatTestResult.Status.FAILED, SatTestResult.Status.BLOCKED]).count()
    remaining = qs.filter(status=SatTestResult.Status.NOT_RUN).count()
    decided = total - remaining
    score = round(100.0 * passed / decided, 1) if decided else None
    return {
        "score": score,
        "total": total,
        "passed": passed,
        "failed": failed,
        "remaining": remaining,
    }


def compute_readiness(run: SatTestRun) -> dict[str, Any]:
    domains = {name: domain_score(run, mods) for name, mods in READINESS_DOMAINS.items()}
    scored = [d["score"] for d in domains.values() if d["score"] is not None]
    overall = round(sum(scored) / len(scored), 1) if scored else None

    open_critical = SatDefect.objects.filter(
        run=run, status=SatDefect.Status.OPEN, severity=SatDefect.Severity.CRITICAL
    ).count()
    open_high = SatDefect.objects.filter(
        run=run, status=SatDefect.Status.OPEN, severity=SatDefect.Severity.HIGH
    ).count()
    failed = SatTestResult.objects.filter(
        run=run, status__in=[SatTestResult.Status.FAILED, SatTestResult.Status.BLOCKED]
    ).count()
    remaining = SatTestResult.objects.filter(run=run, status=SatTestResult.Status.NOT_RUN).count()

    checklist = go_no_go_checklist(run, open_critical=open_critical, open_high=open_high)

    if open_critical > 0:
        recommendation = SatTestRun.Recommendation.NO_GO
    elif remaining > 0:
        recommendation = SatTestRun.Recommendation.PENDING
    elif open_high > 0 or failed > 0:
        recommendation = SatTestRun.Recommendation.CONDITIONAL_GO
    else:
        recommendation = SatTestRun.Recommendation.GO

    return {
        "overall_score": overall,
        "domains": domains,
        "open_critical_defects": open_critical,
        "open_high_defects": open_high,
        "failed_or_blocked_tests": failed,
        "remaining_tests": remaining,
        "recommendation": recommendation,
        "checklist": checklist,
        "as_of": timezone.now().isoformat(),
    }


def go_no_go_checklist(run: SatTestRun, *, open_critical: int, open_high: int) -> list[dict[str, Any]]:
    progress = run_progress(run)
    sat_passed = progress["failed"] == 0 and progress["remaining"] == 0 and progress["passed"] > 0
    return [
        {"id": "phase1", "label": "Phase 1 complete", "passed": True, "notes": "Feature-complete (declared)"},
        {"id": "phase2", "label": "Phase 2 complete", "passed": True, "notes": "Feature-complete (declared)"},
        {"id": "phase25", "label": "Phase 2.5 complete", "passed": True, "notes": "Stabilization complete (declared)"},
        {
            "id": "sat",
            "label": "SAT passed",
            "passed": sat_passed,
            "notes": f"{progress['passed']} passed / {progress['failed']} failed / {progress['remaining']} remaining",
        },
        {
            "id": "uat",
            "label": "UAT passed",
            "passed": False,
            "notes": "Record separately in UAT suite when executed",
        },
        {
            "id": "integration",
            "label": "Integration passed",
            "passed": False,
            "notes": "Record separately in Integration suite when executed",
        },
        {
            "id": "security",
            "label": "Security passed",
            "passed": domain_score(run, READINESS_DOMAINS["Security"])["failed"] == 0
            and domain_score(run, READINESS_DOMAINS["Security"])["remaining"] == 0
            and (domain_score(run, READINESS_DOMAINS["Security"])["passed"] or 0) > 0,
            "notes": "",
        },
        {
            "id": "performance",
            "label": "Performance acceptable",
            "passed": domain_score(run, READINESS_DOMAINS["Performance"])["failed"] == 0
            and domain_score(run, READINESS_DOMAINS["Performance"])["remaining"] == 0,
            "notes": "Confirm measured baselines in Performance Test Plan",
        },
        {"id": "docs", "label": "Documentation complete", "passed": True, "notes": "docs/phase-2.5 + enterprise"},
        {"id": "no_critical", "label": "No Critical defects", "passed": open_critical == 0, "notes": f"{open_critical} open"},
        {"id": "no_high", "label": "No High defects", "passed": open_high == 0, "notes": f"{open_high} open"},
    ]


def live_health_panel() -> dict[str, Any]:
    panel: dict[str, Any] = {
        "as_of": timezone.now().isoformat(),
        "portal": {"status": "online", "detail": "SAT execution API reachable"},
        "highlights": [],
        "nodes": [],
        "alerts": [],
    }
    try:
        alerts = list(LabAlert.objects.filter(status=LabAlert.Status.OPEN).order_by("-severity", "-created_at")[:15])
        panel["alerts"] = [
            {"code": a.code, "severity": a.severity, "title": a.title, "node_id": a.node_id} for a in alerts
        ]
        for a in alerts:
            if a.severity in {LabAlert.Severity.CRITICAL, LabAlert.Severity.ERROR, "critical", "error"}:
                panel["highlights"].append(f"{a.severity.upper()}: {a.title}")
    except Exception as exc:  # noqa: BLE001
        panel["alerts_error"] = str(exc)

    try:
        from iic_booking.lab_infrastructure.services.fleet import build_infrastructure_tree

        tree = build_infrastructure_tree(page=1, page_size=100)
        departments = tree.get("departments") or tree.get("tree") or []
        nodes: list[dict] = []

        def walk(node: dict, trail: list[str]) -> None:
            kind = str(node.get("kind") or "")
            if kind in {"dsa", "equipment_pc", "raa", "analysis_pc"}:
                status = str(node.get("status") or "")
                entry = {
                    "id": node.get("id"),
                    "kind": kind,
                    "name": node.get("computer_name") or node.get("name") or node.get("id"),
                    "status": status,
                    "location_path": " / ".join(trail),
                    "cpu": node.get("cpu"),
                    "memory": node.get("memory"),
                    "disk": node.get("disk"),
                    "last_heartbeat": node.get("last_heartbeat"),
                    "configuration_version": node.get("configuration_version"),
                    "agent_version": node.get("agent_version"),
                    "health_score": node.get("health_score"),
                }
                nodes.append(entry)
                if status.lower() in {"offline", "error", "maintenance"}:
                    panel["highlights"].append(f"{kind} {entry['name']}: {status}")
            name = str(node.get("name") or node.get("department") or node.get("computer_name") or "")
            next_trail = trail + ([name] if name else [])
            for child in node.get("children") or []:
                if isinstance(child, dict):
                    walk(child, next_trail)

        if isinstance(departments, list):
            for d in departments:
                if isinstance(d, dict):
                    walk(d, [])
        panel["nodes"] = nodes[:80]
        panel["dsa"] = {"count": sum(1 for n in nodes if n["kind"] == "dsa")}
        panel["equipment_pc"] = {"count": sum(1 for n in nodes if n["kind"] == "equipment_pc")}
        panel["analysis_pc"] = {"count": sum(1 for n in nodes if n["kind"] in {"raa", "analysis_pc"})}
    except Exception as exc:  # noqa: BLE001
        panel["fleet_error"] = str(exc)
        panel["highlights"].append(f"Fleet snapshot unavailable: {exc}")

    return panel


def build_report_payload(run: SatTestRun) -> dict[str, Any]:
    progress = run_progress(run)
    readiness = run.readiness_snapshot or compute_readiness(run)
    results = (
        SatTestResult.objects.select_related("test_case")
        .filter(run=run)
        .order_by("test_case__stage", "test_case__execution_order")
    )
    defects = SatDefect.objects.filter(run=run).order_by("-created_at")
    evidence = SatEvidence.objects.filter(run=run).order_by("-created_at")
    return {
        "overview": {
            "run_name": run.name,
            "run_id": str(run.id),
            "status": run.status,
            "recommendation": readiness.get("recommendation") or run.recommendation,
            "started_at": progress["started_at"],
            "finished_at": progress["finished_at"],
            "lab_context": run.lab_context or {},
        },
        "summary": progress,
        "stages": stage_summary(run),
        "readiness": readiness,
        "passed_tests": [
            {"test_id": r.test_case.test_id, "feature": r.test_case.feature, "module": r.test_case.module}
            for r in results
            if r.status == SatTestResult.Status.PASSED
        ],
        "failed_tests": [
            {
                "test_id": r.test_case.test_id,
                "feature": r.test_case.feature,
                "actual_result": r.actual_result,
                "failure_snapshot": r.failure_snapshot,
            }
            for r in results
            if r.status == SatTestResult.Status.FAILED
        ],
        "blocked_tests": [
            {"test_id": r.test_case.test_id, "feature": r.test_case.feature, "remarks": r.remarks}
            for r in results
            if r.status == SatTestResult.Status.BLOCKED
        ],
        "evidence": [
            {
                "id": str(e.id),
                "kind": e.kind,
                "title": e.title,
                "original_name": e.original_name,
                "result_id": str(e.result_id) if e.result_id else None,
            }
            for e in evidence[:200]
        ],
        "defects": [
            {
                "id": str(d.id),
                "kind": d.kind,
                "severity": d.severity,
                "status": d.status,
                "title": d.title,
                "test_id": d.test_id,
            }
            for d in defects
        ],
        "recommendations": _recommendations_text(readiness),
        "overall_result": readiness.get("recommendation"),
    }


def _recommendations_text(readiness: dict) -> list[str]:
    rec = readiness.get("recommendation")
    lines = []
    if rec == SatTestRun.Recommendation.GO:
        lines.append("Platform meets SAT gate for production promotion pending formal approval.")
    elif rec == SatTestRun.Recommendation.CONDITIONAL_GO:
        lines.append("Resolve open High defects and re-run failed cases before production.")
    elif rec == SatTestRun.Recommendation.NO_GO:
        lines.append("Do not promote — Critical defects or blocking failures remain.")
    else:
        lines.append("Complete remaining SAT stages and attach evidence before GO decision.")
    if readiness.get("open_critical_defects"):
        lines.append(f"Close {readiness['open_critical_defects']} Critical defect(s).")
    if readiness.get("open_high_defects"):
        lines.append(f"Close {readiness['open_high_defects']} High defect(s).")
    if readiness.get("remaining_tests"):
        lines.append(f"Execute {readiness['remaining_tests']} remaining test(s).")
    return lines


def report_csv(run: SatTestRun) -> HttpResponse:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(
        ["test_id", "stage", "module", "feature", "severity", "status", "actual_result", "remarks", "executed_at"]
    )
    for r in SatTestResult.objects.select_related("test_case").filter(run=run).order_by(
        "test_case__stage", "test_case__execution_order"
    ):
        writer.writerow(
            [
                r.test_case.test_id,
                r.test_case.stage,
                r.test_case.module,
                r.test_case.feature,
                r.test_case.severity,
                r.status,
                r.actual_result,
                r.remarks,
                r.executed_at.isoformat() if r.executed_at else "",
            ]
        )
    resp = HttpResponse(buf.getvalue(), content_type="text/csv")
    resp["Content-Disposition"] = f'attachment; filename="sat-report-{run.id}.csv"'
    return resp


def report_xlsx(run: SatTestRun) -> HttpResponse:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    ws.append(["test_id", "stage", "module", "feature", "severity", "status", "actual_result", "remarks"])
    for r in SatTestResult.objects.select_related("test_case").filter(run=run).order_by(
        "test_case__stage", "test_case__execution_order"
    ):
        ws.append(
            [
                r.test_case.test_id,
                r.test_case.stage,
                r.test_case.module,
                r.test_case.feature,
                r.test_case.severity,
                r.status,
                r.actual_result,
                r.remarks,
            ]
        )
    summary = wb.create_sheet("Summary")
    payload = build_report_payload(run)
    summary.append(["Field", "Value"])
    for k, v in payload["summary"].items():
        summary.append([k, json.dumps(v) if isinstance(v, (dict, list)) else v])
    summary.append([])
    summary.append(["Domain", "Score"])
    for name, dom in (payload["readiness"].get("domains") or {}).items():
        summary.append([name, dom.get("score")])
    defects = wb.create_sheet("Defects")
    defects.append(["id", "kind", "severity", "status", "title", "test_id"])
    for d in payload["defects"]:
        defects.append([d["id"], d["kind"], d["severity"], d["status"], d["title"], d["test_id"]])
    out = io.BytesIO()
    wb.save(out)
    resp = HttpResponse(
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="sat-report-{run.id}.xlsx"'
    return resp


def report_pdf(run: SatTestRun) -> HttpResponse:
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas

    payload = build_report_payload(run)
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    width, height = A4
    y = height - 40
    c.setFont("Helvetica-Bold", 14)
    c.drawString(40, y, f"SAT Report — {run.name}")
    y -= 24
    c.setFont("Helvetica", 10)
    lines = [
        f"Run ID: {run.id}",
        f"Status: {run.status}",
        f"Recommendation: {payload.get('overall_result')}",
        f"Passed: {payload['summary']['passed']}  Failed: {payload['summary']['failed']}  "
        f"Blocked: {payload['summary']['blocked']}  Remaining: {payload['summary']['remaining']}",
        f"Completion: {payload['summary']['completion_pct']}%",
        "",
        "Recommendations:",
        *payload.get("recommendations", []),
        "",
        "Failed tests:",
    ]
    for t in payload.get("failed_tests", [])[:30]:
        lines.append(f"- {t['test_id']}: {t.get('feature', '')}")
    for line in lines:
        if y < 50:
            c.showPage()
            y = height - 40
            c.setFont("Helvetica", 10)
        c.drawString(40, y, str(line)[:110])
        y -= 14
    c.save()
    resp = HttpResponse(buf.getvalue(), content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="sat-report-{run.id}.pdf"'
    return resp
