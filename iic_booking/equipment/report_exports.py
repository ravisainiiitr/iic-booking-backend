"""
Export equipment report to PDF and Excel.
"""

import io
import tempfile
from typing import Any, Optional

from .reports import get_equipment_report_data


def _report_duration_caption(data: dict) -> str:
    """Human-readable duration line for PDF/Excel (matches API report_header)."""
    hdr = data.get("report_header") or {}
    human = hdr.get("period_display")
    if not human:
        df = data.get("date_from") or ""
        dt = data.get("date_to") or ""
        human = f"{df} to {dt}" if df and dt else (df or dt or "—")
    sfx = str(hdr.get("report_duration_suffix") or "")
    return f"Report Duration: {human}{sfx}"


def build_report_pdf(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    equipment_ids: Optional[list[int]] = None,
) -> bytes:
    """Build equipment report as PDF bytes."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
        PageBreak,
    )
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    data = get_equipment_report_data(
        date_from=date_from,
        date_to=date_to,
        equipment_ids=equipment_ids,
    )
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        name="ReportTitle",
        parent=styles["Heading1"],
        fontSize=16,
        spaceAfter=12,
        alignment=TA_CENTER,
    )
    heading_style = ParagraphStyle(
        name="SectionHeading",
        parent=styles["Heading2"],
        fontSize=12,
        spaceAfter=6,
        spaceBefore=12,
    )
    body_style = styles["Normal"]
    small_style = ParagraphStyle(
        name="ReportSmall",
        parent=body_style,
        fontSize=8,
        leading=10,
    )
    subtitle_style = ParagraphStyle(
        name="ReportSubtitle",
        parent=body_style,
        fontSize=10,
        alignment=TA_CENTER,
        spaceAfter=8,
        textColor=colors.HexColor("#374151"),
    )

    hdr = data.get("report_header") or {}
    story = []
    dept_name = str(hdr.get("department_name") or hdr.get("institute_name") or "—")
    story.append(Paragraph(dept_name, title_style))
    story.append(Paragraph("Indian Institute of Technology Roorkee", subtitle_style))
    story.append(Paragraph(str(hdr.get("report_title", "Equipment Performance Report")), title_style))
    story.append(Paragraph(_report_duration_caption(data), subtitle_style))
    story.append(Spacer(1, 0.4 * cm))

    # Financial + KPI summary
    story.append(Paragraph("Financial & KPI summary", heading_style))
    summary = data.get("summary", {}) or {}
    fin_rows = [
        ["Metric", "Value"],
        ["Revenue (total)", f"₹{float(summary.get('revenue_total', 0) or 0):.2f}"],
        ["Revenue (internal)", f"₹{float(summary.get('revenue_internal', 0) or 0):.2f}"],
        ["Revenue (external)", f"₹{float(summary.get('revenue_external', 0) or 0):.2f}"],
        ["Slot hours (all statuses)", f"{float(summary.get('total_hours', 0) or 0):.2f}"],
        ["Utilized hours (BOOKED slots)", f"{float(summary.get('utilized_hours', 0) or 0):.2f}"],
        ["Downtime hours (maint. + op. absent)", f"{float(summary.get('downtime_hours', 0) or 0):.2f}"],
        ["Utilization factor (booked / all slot hours)", f"{float(summary.get('utilization_factor', 0) or 0) * 100:.2f}%"],
        [
            "Available hours (Mon–Fri, excl. holidays; slot time window)",
            f"{float(summary.get('available_hours_working_window', 0) or 0):.2f}",
        ],
        [
            "Completed booking hours (same window)",
            f"{float(summary.get('completed_hours_in_working_window', 0) or 0):.2f}",
        ],
        [
            "Utilization vs working capacity (completed / available window)",
            f"{float(summary.get('utilization_vs_working_capacity', 0) or 0) * 100:.2f}%",
        ],
    ]
    t0 = Table(fin_rows, repeatRows=1)
    t0.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
        ])
    )
    story.append(t0)
    story.append(Spacer(1, 0.6 * cm))

    # Per-equipment performance (two-column metric tables)
    story.append(Paragraph("Per-equipment performance", heading_style))
    crit_labels = {
        "on_time_operator_availability": "On-time & operator availability (Yes)",
        "laboratory_cleanliness_organization": "Laboratory cleanliness & organization (Yes)",
        "sample_handling_care": "Sample handling & care (Yes)",
        "operator_behaviour_professionalism": "Operator behaviour & professionalism (Yes)",
        "compliance_booking_request_parameters": "Compliance with booking parameters (Yes)",
    }

    for idx, eq in enumerate(data.get("equipment") or []):
        if idx > 0:
            story.append(PageBreak())
        oics = eq.get("officers_in_charge") or []
        ops = eq.get("lab_operators") or []
        oic_line = ", ".join(f"{x.get('name', '')} ({x.get('email', '')})" for x in oics) or "—"
        op_line = ", ".join(f"{x.get('name', '')} ({x.get('email', '')})" for x in ops) or "—"
        story.append(
            Paragraph(
                f"<b>{eq.get('name', '')}</b> &nbsp;({eq.get('code', '')})",
                ParagraphStyle(
                    name=f"EqTitle_{idx}",
                    parent=heading_style,
                    fontSize=11,
                    spaceBefore=4,
                    textColor=colors.HexColor("#1e3a5f"),
                ),
            )
        )
        story.append(Paragraph(f"<b>Officer(s) in charge:</b> {oic_line}", small_style))
        story.append(Paragraph(f"<b>Lab operator(s):</b> {op_line}", small_style))
        story.append(Paragraph(f"<b>Slot time window:</b> {eq.get('slot_window_display', '')}", small_style))
        story.append(Spacer(1, 0.25 * cm))

        perf_rows = [
            ["Metric", "Value"],
            ["Distinct users served (excl. cancelled/refunded/waitlist/pending)", str(eq.get("distinct_users_served", 0))],
            ["Distinct internal users", str(eq.get("distinct_users_internal", 0))],
            ["Distinct external users", str(eq.get("distinct_users_external", 0))],
            ["Total samples (input A)", str(eq.get("total_samples", 0))],
            ["Samples — internal / external", f"{eq.get('samples_internal', 0)} / {eq.get('samples_external', 0)}"],
            [
                "Booking hours (total / int. / ext.)",
                f"{eq.get('total_booking_hours', 0)} / {eq.get('booking_hours_internal', 0)} / {eq.get('booking_hours_external', 0)}",
            ],
            [
                "Available hours (working window, Mon–Fri excl. holidays)",
                str(eq.get("available_hours_working_window", 0)),
            ],
            [
                "Slot hours on weekends & holidays (incl. institute holidays)",
                str(eq.get("available_hours_weekend_or_holiday", 0)),
            ],
            [
                "Completed booking hours (within working window)",
                str(eq.get("completed_slot_hours_working_window", 0)),
            ],
            [
                "Utilization vs working capacity",
                f"{float(eq.get('utilization_vs_working_capacity', 0) or 0) * 100:.2f}%",
            ],
            ["Hours — booking not utilized", str(eq.get("booking_not_utilized_hours", 0))],
            ["Hours — under maintenance", str(eq.get("under_maintenance_hours", 0))],
            ["Hours — operator absent", str(eq.get("operator_absent_hours", 0))],
            ["Hours — other disruption (linked bookings)", str(eq.get("other_disruption_hours", 0))],
            ["Hours — blocked slots", str(eq.get("blocked_hours", 0))],
            ["Bookings in period / completed", f"{eq.get('total_bookings_in_period', 0)} / {eq.get('completed_in_period', 0)}"],
        ]
        tp = Table(perf_rows, colWidths=[10 * cm, 7 * cm], repeatRows=1)
        tp.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f766e")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (0, -1), "LEFT"),
                ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#f8fafc")),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#99f6e4")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ])
        )
        story.append(tp)
        story.append(Spacer(1, 0.35 * cm))

        ur = eq.get("user_ratings") or {}
        n_r = int(ur.get("ratings_submitted_count", 0) or 0)
        story.append(Paragraph(f"<b>User ratings</b> (submitted in period: {n_r})", small_style))
        if n_r > 0:
            rrows = [["Criterion", "Yes", "No", "Unanswered"]]
            for key, label in crit_labels.items():
                c = ur.get("criteria", {}).get(key, {})
                rrows.append([
                    label,
                    str(c.get("yes", 0)),
                    str(c.get("no", 0)),
                    str(c.get("unanswered", 0)),
                ])
            avg = ur.get("overall_rating_avg")
            rrows.append(["Overall rating (0–5) avg.", str(avg) if avg is not None else "—", "", ""])
            tr = Table(rrows, colWidths=[8 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm], repeatRows=1)
            tr.setStyle(
                TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4c1d95")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                    ("ALIGN", (1, 0), (-1, -1), "CENTER"),
                ])
            )
            story.append(tr)
        else:
            story.append(Paragraph("No ratings submitted in this period.", small_style))
        story.append(Spacer(1, 0.5 * cm))

    # Revenue breakdown tables (top rows)
    financial = data.get("financial", {}) or {}
    r_equipment = financial.get("revenue_by_equipment", []) or []
    if r_equipment:
        story.append(Paragraph("Revenue by equipment (top 20)", heading_style))
        rows_rev_eq = [["Equipment", "Bookings", "Revenue"]]
        for r in r_equipment[:20]:
            rows_rev_eq.append([
                f"{r.get('equipment__code','')} — {r.get('equipment__name','')}",
                str(r.get("count", 0) or 0),
                f"₹{float(r.get('total', 0) or 0):.2f}",
            ])
        tt = Table(rows_rev_eq, repeatRows=1)
        tt.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0ea5e9")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
            ])
        )
        story.append(tt)
        story.append(Spacer(1, 0.4 * cm))

    r_ext = financial.get("revenue_by_external_category", []) or []
    if r_ext:
        story.append(Paragraph("External revenue by category", heading_style))
        rows_rev_ext = [["Category", "Bookings", "Revenue"]]
        for r in r_ext:
            rows_rev_ext.append([
                str(r.get("user_type_snapshot", "") or ""),
                str(r.get("count", 0) or 0),
                f"₹{float(r.get('total', 0) or 0):.2f}",
            ])
        tt2 = Table(rows_rev_ext, repeatRows=1)
        tt2.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#9333ea")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
            ])
        )
        story.append(tt2)
        story.append(Spacer(1, 0.4 * cm))

    # Per-equipment utilization pie charts
    story.append(Paragraph("Utilization by equipment (hours)", heading_style))
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.piecharts import Pie as PieChartDrawing

    UTILIZATION_CATEGORIES = [
        ("booked_hours", "Utilized (Booked)"),
        ("booking_not_utilized_hours", "Booking not utilized"),
        ("under_maintenance_hours", "Under maintenance"),
        ("operator_absent_hours", "Operator absent"),
        ("other_disruption_hours", "Other disruption"),
        ("blocked_hours", "Blocked"),
        ("no_booking_hours", "No booking"),
    ]
    PIE_COLORS_PDF = [
        colors.HexColor("#22c55e"),
        colors.HexColor("#a855f7"),
        colors.HexColor("#f97316"),
        colors.HexColor("#eab308"),
        colors.HexColor("#dc2626"),
        colors.HexColor("#78716c"),
        colors.HexColor("#64748b"),
    ]
    for pie_idx, eq in enumerate(data["equipment"]):
        name = eq.get("name", "") or "Equipment"
        code = eq.get("code", "")
        eq_heading_style = ParagraphStyle(
            name=f"EqPieHeading_{pie_idx}",
            parent=styles["Normal"],
            fontSize=10,
            fontName="Helvetica-Bold",
            spaceBefore=8,
            spaceAfter=4,
        )
        story.append(Paragraph(f"{name} ({code})", eq_heading_style))
        values = []
        labels = []
        for i, (key, label) in enumerate(UTILIZATION_CATEGORIES):
            h = float(eq.get(key, 0) or 0)
            if h > 0:
                values.append(h)
                labels.append(label)
        if not values:
            story.append(Paragraph("No slot data in period", body_style))
            story.append(Spacer(1, 0.2 * cm))
            continue
        drawing = Drawing(220, 140)
        pie = PieChartDrawing()
        pie.x = 10
        pie.y = 10
        pie.width = 120
        pie.height = 120
        pie.data = values
        pie.labels = labels
        pie.slices.strokeWidth = 0.5
        for i, _ in enumerate(values):
            pie.slices[i].fillColor = PIE_COLORS_PDF[i % len(PIE_COLORS_PDF)]
        drawing.add(pie)
        story.append(drawing)
        story.append(Spacer(1, 0.3 * cm))

    story.append(Spacer(1, 0.5 * cm))

    story.append(Paragraph("Overall utilization (all equipment, hours)", heading_style))
    pie_rows = [["Category", "Hours"]]
    for p in data.get("utilization_pie", []):
        pie_rows.append([str(p.get("name", "")), str(p.get("hours", 0))])
    t2 = Table(pie_rows, repeatRows=1)
    t2.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#70AD47")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ])
    )
    story.append(t2)

    doc.build(story)
    return buffer.getvalue()


def build_report_excel(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    equipment_ids: Optional[list[int]] = None,
) -> bytes:
    """Build equipment report as Excel (xlsx) bytes."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    data = get_equipment_report_data(
        date_from=date_from,
        date_to=date_to,
        equipment_ids=equipment_ids,
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Equipment Report"
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    hdr = data.get("report_header") or {}
    ws["A1"] = "Institute Instrumentation Centre — IIT Roorkee"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = str(hdr.get("report_title", "Equipment Performance Report"))
    ws["A2"].font = Font(bold=True, size=11)
    ws["A3"] = _report_duration_caption(data)
    row = 5

    # Financial summary
    summary = data.get("summary", {}) or {}
    ws.cell(row=row, column=1, value="Financial & KPI summary").font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value="Revenue (total)")
    ws.cell(row=row, column=2, value=float(summary.get("revenue_total", 0) or 0))
    row += 1
    ws.cell(row=row, column=1, value="Revenue (internal)")
    ws.cell(row=row, column=2, value=float(summary.get("revenue_internal", 0) or 0))
    row += 1
    ws.cell(row=row, column=1, value="Revenue (external)")
    ws.cell(row=row, column=2, value=float(summary.get("revenue_external", 0) or 0))
    row += 1
    ws.cell(row=row, column=1, value="Total hours")
    ws.cell(row=row, column=2, value=float(summary.get("total_hours", 0) or 0))
    row += 1
    ws.cell(row=row, column=1, value="Utilized hours")
    ws.cell(row=row, column=2, value=float(summary.get("utilized_hours", 0) or 0))
    row += 1
    ws.cell(row=row, column=1, value="Downtime hours")
    ws.cell(row=row, column=2, value=float(summary.get("downtime_hours", 0) or 0))
    row += 1
    ws.cell(row=row, column=1, value="Utilization factor (booked / all slot hours)")
    ws.cell(row=row, column=2, value=float(summary.get("utilization_factor", 0) or 0))
    row += 1
    ws.cell(row=row, column=1, value="Available hours (working window)")
    ws.cell(row=row, column=2, value=float(summary.get("available_hours_working_window", 0) or 0))
    row += 1
    ws.cell(row=row, column=1, value="Completed hours (working window)")
    ws.cell(row=row, column=2, value=float(summary.get("completed_hours_in_working_window", 0) or 0))
    row += 1
    ws.cell(row=row, column=1, value="Utilization vs working capacity")
    ws.cell(row=row, column=2, value=float(summary.get("utilization_vs_working_capacity", 0) or 0))
    row += 2

    # Per-equipment headers
    headers = [
        "Equipment",
        "Code",
        "Status",
        "OIC names",
        "Lab operators",
        "Slot window",
        "Distinct users",
        "Users int.",
        "Users ext.",
        "Samples (A)",
        "Samples int.",
        "Samples ext.",
        "Book hrs total",
        "Book hrs int.",
        "Book hrs ext.",
        "Avail. work hrs",
        "Wknd/hol hrs",
        "Compl. work hrs",
        "Util. vs cap.",
        "Bookings period",
        "Completed",
        "Overall bookings",
        "Current BOOKED",
        "Under maint. slots",
        "Under maint. hrs",
        "Op. absent slots",
        "Op. absent hrs",
        "Not util. slots",
        "Not util. hrs",
        "No booking slots",
        "No booking hrs",
        "Booked slots",
        "Booked hrs",
        "Blocked hrs",
        "Other disrupt. hrs",
        "Ratings count",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border
    row += 1
    for eq in data["equipment"]:
        oics = ", ".join(str(x.get("name", "")) for x in (eq.get("officers_in_charge") or []))
        ops = ", ".join(str(x.get("name", "")) for x in (eq.get("lab_operators") or []))
        ur = eq.get("user_ratings") or {}
        col = 1
        ws.cell(row=row, column=col, value=eq.get("name", ""))
        col += 1
        ws.cell(row=row, column=col, value=eq.get("code", ""))
        col += 1
        ws.cell(row=row, column=col, value=eq.get("status_display", "") or eq.get("status", "") or "")
        col += 1
        ws.cell(row=row, column=col, value=oics)
        col += 1
        ws.cell(row=row, column=col, value=ops)
        col += 1
        ws.cell(row=row, column=col, value=eq.get("slot_window_display", ""))
        col += 1
        ws.cell(row=row, column=col, value=eq.get("distinct_users_served", 0))
        col += 1
        ws.cell(row=row, column=col, value=eq.get("distinct_users_internal", 0))
        col += 1
        ws.cell(row=row, column=col, value=eq.get("distinct_users_external", 0))
        col += 1
        ws.cell(row=row, column=col, value=eq.get("total_samples", 0))
        col += 1
        ws.cell(row=row, column=col, value=eq.get("samples_internal", 0))
        col += 1
        ws.cell(row=row, column=col, value=eq.get("samples_external", 0))
        col += 1
        ws.cell(row=row, column=col, value=eq.get("total_booking_hours", 0))
        col += 1
        ws.cell(row=row, column=col, value=eq.get("booking_hours_internal", 0))
        col += 1
        ws.cell(row=row, column=col, value=eq.get("booking_hours_external", 0))
        col += 1
        ws.cell(row=row, column=col, value=eq.get("available_hours_working_window", 0))
        col += 1
        ws.cell(row=row, column=col, value=eq.get("available_hours_weekend_or_holiday", 0))
        col += 1
        ws.cell(row=row, column=col, value=eq.get("completed_slot_hours_working_window", 0))
        col += 1
        ws.cell(row=row, column=col, value=float(eq.get("utilization_vs_working_capacity", 0) or 0))
        col += 1
        ws.cell(row=row, column=col, value=eq.get("total_bookings_in_period", 0))
        col += 1
        ws.cell(row=row, column=col, value=eq.get("completed_in_period", 0))
        col += 1
        ws.cell(row=row, column=col, value=eq.get("overall_bookings", 0))
        col += 1
        ws.cell(row=row, column=col, value=eq.get("overall_current_bookings", 0))
        col += 1
        ws.cell(row=row, column=col, value=eq.get("under_maintenance_slots", 0))
        col += 1
        ws.cell(row=row, column=col, value=eq.get("under_maintenance_hours", 0))
        col += 1
        ws.cell(row=row, column=col, value=eq.get("operator_absent_slots", 0))
        col += 1
        ws.cell(row=row, column=col, value=eq.get("operator_absent_hours", 0))
        col += 1
        ws.cell(row=row, column=col, value=eq.get("booking_not_utilized_slots", 0))
        col += 1
        ws.cell(row=row, column=col, value=eq.get("booking_not_utilized_hours", 0))
        col += 1
        ws.cell(row=row, column=col, value=eq.get("no_booking_slots", 0))
        col += 1
        ws.cell(row=row, column=col, value=eq.get("no_booking_hours", 0))
        col += 1
        ws.cell(row=row, column=col, value=eq.get("booked_slots", 0))
        col += 1
        ws.cell(row=row, column=col, value=eq.get("booked_hours", 0))
        col += 1
        ws.cell(row=row, column=col, value=eq.get("blocked_hours", 0))
        col += 1
        ws.cell(row=row, column=col, value=eq.get("other_disruption_hours", 0))
        col += 1
        ws.cell(row=row, column=col, value=ur.get("ratings_submitted_count", 0))
        for c in range(1, len(headers) + 1):
            ws.cell(row=row, column=c).border = thin_border
        row += 1
    row += 1
    ws.cell(row=row, column=1, value="Overall utilization (hours)")
    ws.cell(row=row, column=1).font = header_font
    row += 1
    ws.cell(row=row, column=1, value="Category")
    ws.cell(row=row, column=2, value="Hours")
    for c in range(1, 3):
        ws.cell(row=row, column=c).font = header_font
        ws.cell(row=row, column=c).border = thin_border
    row += 1
    for p in data.get("utilization_pie", []):
        ws.cell(row=row, column=1, value=p.get("name", ""))
        ws.cell(row=row, column=2, value=p.get("hours", 0))
        for c in range(1, 3):
            ws.cell(row=row, column=c).border = thin_border
        row += 1

    # Revenue breakdown sheets
    financial = data.get("financial", {}) or {}
    r_equipment = financial.get("revenue_by_equipment", []) or []
    if r_equipment:
        ws3 = wb.create_sheet("Revenue by Equipment", 2)
        ws3["A1"] = "Revenue by Equipment (completed bookings)"
        ws3["A1"].font = Font(bold=True, size=12)
        ws3["A2"] = _report_duration_caption(data)
        rr = 4
        ws3.cell(row=rr, column=1, value="Equipment")
        ws3.cell(row=rr, column=2, value="Bookings")
        ws3.cell(row=rr, column=3, value="Revenue")
        for c in range(1, 4):
            ws3.cell(row=rr, column=c).font = header_font
            ws3.cell(row=rr, column=c).border = thin_border
        rr += 1
        for r in r_equipment:
            ws3.cell(row=rr, column=1, value=f"{r.get('equipment__code','')} — {r.get('equipment__name','')}")
            ws3.cell(row=rr, column=2, value=int(r.get("count", 0) or 0))
            ws3.cell(row=rr, column=3, value=float(r.get("total", 0) or 0))
            for c in range(1, 4):
                ws3.cell(row=rr, column=c).border = thin_border
            rr += 1

    r_ext = financial.get("revenue_by_external_category", []) or []
    if r_ext:
        ws4 = wb.create_sheet("External Revenue", 3)
        ws4["A1"] = "External Revenue by Category (completed bookings)"
        ws4["A1"].font = Font(bold=True, size=12)
        ws4["A2"] = _report_duration_caption(data)
        rr = 4
        ws4.cell(row=rr, column=1, value="Category")
        ws4.cell(row=rr, column=2, value="Bookings")
        ws4.cell(row=rr, column=3, value="Revenue")
        for c in range(1, 4):
            ws4.cell(row=rr, column=c).font = header_font
            ws4.cell(row=rr, column=c).border = thin_border
        rr += 1
        for r in r_ext:
            ws4.cell(row=rr, column=1, value=str(r.get("user_type_snapshot", "") or ""))
            ws4.cell(row=rr, column=2, value=int(r.get("count", 0) or 0))
            ws4.cell(row=rr, column=3, value=float(r.get("total", 0) or 0))
            for c in range(1, 4):
                ws4.cell(row=rr, column=c).border = thin_border
            rr += 1

    # Per-equipment utilization: new sheet with pie table + chart per equipment
    from openpyxl.chart import PieChart, Reference

    ws2 = wb.create_sheet("Per-Equipment Utilization", 1)
    ws2["A1"] = "Per-Equipment Utilization (hours)"
    ws2["A1"].font = Font(bold=True, size=12)
    ws2["A2"] = _report_duration_caption(data)
    eq_row = 4
    UTILIZATION_KEYS = [
        ("booked_hours", "Utilized (Booked)"),
        ("booking_not_utilized_hours", "Booking not utilized"),
        ("under_maintenance_hours", "Under maintenance"),
        ("operator_absent_hours", "Operator absent"),
        ("other_disruption_hours", "Other disruption"),
        ("blocked_hours", "Blocked"),
        ("no_booking_hours", "No booking"),
    ]
    for eq in data["equipment"]:
        name = eq.get("name", "") or "Equipment"
        code = eq.get("code", "")
        ws2.cell(row=eq_row, column=1, value=f"{name} ({code})")
        ws2.cell(row=eq_row, column=1).font = Font(bold=True)
        eq_row += 1
        ws2.cell(row=eq_row, column=1, value="Category")
        ws2.cell(row=eq_row, column=2, value="Hours")
        for c in range(1, 3):
            ws2.cell(row=eq_row, column=c).font = header_font
            ws2.cell(row=eq_row, column=c).border = thin_border
        eq_row += 1
        data_start = eq_row
        for key, label in UTILIZATION_KEYS:
            h = float(eq.get(key, 0) or 0)
            ws2.cell(row=eq_row, column=1, value=label)
            ws2.cell(row=eq_row, column=2, value=h)
            for c in range(1, 3):
                ws2.cell(row=eq_row, column=c).border = thin_border
            eq_row += 1
        data_end = eq_row - 1
        has_any_hours = any(float(eq.get(key, 0) or 0) > 0 for key, _ in UTILIZATION_KEYS)
        if has_any_hours and data_end >= data_start:
            chart = PieChart()
            chart.title = f"{code} utilization"
            chart.width = 12
            chart.height = 8
            data_ref = Reference(ws2, min_col=2, min_row=data_start, max_row=data_end)
            labels_ref = Reference(ws2, min_col=1, min_row=data_start, max_row=data_end)
            chart.add_data(data_ref, titles_from_data=False)
            chart.set_categories(labels_ref)
            ws2.add_chart(chart, f"D{data_start}")
        eq_row += 2

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
